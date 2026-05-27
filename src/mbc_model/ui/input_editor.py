"""Validated input workbook editing for the desktop UI."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from mbc_model.data.loader import ExcelLoader

MAX_INFORCE_EDIT_ROWS = 100
MAX_RANDOM_EDIT_SIZE = 99

_INFORCE_FIELDS = {
    "YOB": "YOB",
    "Gender": "Gender",
    "Annual Benefit": "Annual Benefit",
}
_PARAMETER_SCALARS = {
    "Valuation Date",
    "Last Projection Year",
    "Which Random Numbers?",
    "Random Numbers Seed",
}
_REPORTING_FIELDS = {
    "Policy Results": {"Create?", "Policy", "Scenario"},
    "Scenario Results": {"Create?", "Scenario", "Create graph?"},
    "Total Results": {"Create?"},
    "Dashboard Results": {
        "Create?",
        "Scenarios",
        "Policies",
        "Discount Rate",
        "Create graph?",
    },
}


@dataclass(frozen=True)
class ValidationResult:
    """Validation result returned before Save As is opened."""

    ok: bool
    errors: list[str]


class InputEditError(ValueError):
    """Raised when edits cannot be applied safely."""


def validate_input_changes(input_path: Path, changes: dict[str, Any]) -> ValidationResult:
    """Validate a UI edit payload without saving a workbook."""
    errors = _validate_changes(input_path, changes)
    return ValidationResult(ok=not errors, errors=errors)


def save_input_copy(input_path: Path, output_path: Path, changes: dict[str, Any]) -> Path:
    """Save a copy of an input workbook with validated UI edits applied."""
    input_path = input_path.resolve()
    output_path = output_path.resolve()
    if input_path == output_path:
        raise InputEditError("Save As must use a different file name from the source workbook.")
    if output_path.suffix.lower() != ".xlsx":
        raise InputEditError("Edited input workbooks must be saved as .xlsx files.")

    errors = _validate_changes(input_path, changes)
    if errors:
        raise InputEditError(errors[0])

    workbook = openpyxl.load_workbook(input_path, keep_vba=False)
    _apply_inforce_changes(workbook, input_path, changes.get("inforce", []))
    _apply_parameter_changes(workbook, input_path, changes.get("parameters", {}))
    _apply_reporting_changes(workbook, changes.get("reporting", []))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _validate_changes(input_path: Path, changes: dict[str, Any]) -> list[str]:
    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=False, keep_vba=True)
    data_workbook = openpyxl.load_workbook(
        input_path, read_only=True, data_only=True, keep_vba=True
    )
    errors: list[str] = []

    try:
        _validate_inforce_changes(workbook, input_path, changes.get("inforce", []), errors)
        _validate_parameter_changes(
            workbook,
            data_workbook,
            input_path,
            changes.get("parameters", {}),
            errors,
        )
        _validate_reporting_changes(workbook, changes.get("reporting", []), errors)
    finally:
        workbook.close()
        data_workbook.close()

    return errors


def _validate_inforce_changes(
    workbook: Any,
    input_path: Path,
    edits: list[dict[str, Any]],
    errors: list[str],
) -> None:
    if not edits:
        return
    worksheet = workbook["Inforce"]
    policy_rows = _inforce_policy_rows(worksheet)
    if len(policy_rows) > MAX_INFORCE_EDIT_ROWS:
        errors.append("Inforce edits can only be saved for workbooks with 100 policies or fewer.")
        return

    columns = _header_columns(worksheet, 1)
    for edit in edits:
        policy_id = _int_value(edit.get("policy_id"), "Inforce policy", errors)
        field = str(edit.get("field", "")).strip()
        if policy_id is None:
            continue
        if policy_id not in policy_rows:
            errors.append(f"Inforce policy {policy_id} is not in the workbook.")
            continue
        if field not in _INFORCE_FIELDS:
            errors.append(f"Inforce field {field!r} is not editable.")
            continue
        row = policy_rows[policy_id]
        column = columns.get(_INFORCE_FIELDS[field])
        if column is None:
            errors.append(f"Inforce column {field!r} was not found.")
            continue
        _reject_formula(worksheet.cell(row, column), f"Inforce policy {policy_id} {field}", errors)
        _coerce_inforce_value(field, edit.get("value"), errors)


def _validate_parameter_changes(
    workbook: Any,
    data_workbook: Any,
    input_path: Path,
    edits: dict[str, Any],
    errors: list[str],
) -> None:
    if not edits:
        return
    worksheet = workbook["Parameters"]
    data_worksheet = data_workbook["Parameters"]

    for edit in edits.get("projection_settings", []):
        parameter = str(edit.get("parameter", "")).strip()
        if parameter not in _PARAMETER_SCALARS:
            errors.append(f"Parameter {parameter!r} is not editable.")
            continue
        row = _find_marker_row(worksheet, parameter)
        if row is None:
            errors.append(f"Parameter {parameter!r} was not found.")
            continue
        _reject_formula(worksheet.cell(row, 3), parameter, errors)
        _coerce_parameter_scalar(parameter, edit.get("value"), errors)

    _validate_gender_table_edits(
        worksheet,
        "Mortality Rates",
        edits.get("mortality_rates", []),
        errors,
    )
    _validate_gender_table_edits(
        worksheet,
        "Projection Scale",
        edits.get("projection_scale", []),
        errors,
    )
    _validate_random_number_edits(
        worksheet,
        data_worksheet,
        input_path,
        edits.get("random_numbers", []),
        errors,
    )


def _validate_gender_table_edits(
    worksheet: Any,
    marker: str,
    edits: list[dict[str, Any]],
    errors: list[str],
) -> None:
    if not edits:
        return
    start_row = _find_marker_row(worksheet, marker)
    if start_row is None:
        errors.append(f"{marker} table was not found.")
        return
    for edit in edits:
        age = _int_value(edit.get("age"), f"{marker} age", errors)
        if age is None:
            continue
        row = start_row + 1 + age
        if row > worksheet.max_row:
            errors.append(f"{marker} age {age} is outside the workbook table.")
            continue
        for field, column in (("Male", 4), ("Female", 5)):
            if field.lower() not in edit:
                continue
            label = f"{marker} age {age} {field}"
            _reject_formula(worksheet.cell(row, column), label, errors)
            _rate_value(edit.get(field.lower()), label, errors)


def _validate_random_number_edits(
    worksheet: Any,
    data_worksheet: Any,
    input_path: Path,
    edits: list[dict[str, Any]],
    errors: list[str],
) -> None:
    if not edits:
        return
    policies = ExcelLoader(input_path).load_inforce()
    _, _, random_table = ExcelLoader(input_path).load_parameters()
    if random_table is None:
        errors.append(
            "Random Numbers can only be edited when the workbook uses a random number table."
        )
        return
    scenario_count, policy_count = int(random_table.shape[0]), len(policies)
    if scenario_count >= MAX_RANDOM_EDIT_SIZE + 1 or policy_count >= MAX_RANDOM_EDIT_SIZE + 1:
        errors.append(
            "Random Numbers are editable only when scenarios and policies are both under 100."
        )
        return

    start_row = _find_marker_row(data_worksheet, "Random Numbers")
    if start_row is None:
        errors.append("Random Numbers table was not found.")
        return
    for edit in edits:
        scenario = _int_value(edit.get("scenario"), "Random Numbers scenario", errors)
        policy = _int_value(edit.get("policy"), "Random Numbers policy", errors)
        if scenario is None or policy is None:
            continue
        if not (1 <= scenario <= scenario_count and 1 <= policy <= policy_count):
            errors.append(
                f"Random Numbers scenario {scenario}, policy {policy} is outside the table."
            )
            continue
        row = start_row + scenario
        column = 3 + policy
        _reject_formula(worksheet.cell(row, column), f"Random Numbers {scenario}/{policy}", errors)
        _rate_value(
            edit.get("value"), f"Random Numbers scenario {scenario}, policy {policy}", errors
        )


def _validate_reporting_changes(
    workbook: Any,
    edits: list[dict[str, Any]],
    errors: list[str],
) -> None:
    if not edits:
        return
    worksheet = workbook["Reporting"]
    for edit in edits:
        section = str(edit.get("section", "")).strip()
        field = str(edit.get("field", "")).strip()
        if section not in _REPORTING_FIELDS or field not in _REPORTING_FIELDS[section]:
            errors.append(f"Reporting field {section} / {field} is not editable.")
            continue
        row = _find_reporting_row(worksheet, section, field)
        if row is None:
            errors.append(f"Reporting field {section} / {field} was not found.")
            continue
        _reject_formula(worksheet.cell(row, 3), f"Reporting {section} {field}", errors)
        _coerce_reporting_value(section, field, edit.get("value"), errors)


def _apply_inforce_changes(workbook: Any, input_path: Path, edits: list[dict[str, Any]]) -> None:
    if not edits:
        return
    worksheet = workbook["Inforce"]
    policy_rows = _inforce_policy_rows(worksheet)
    columns = _header_columns(worksheet, 1)
    for edit in edits:
        policy_id = int(edit["policy_id"])
        if policy_id not in policy_rows:
            continue
        field = str(edit["field"]).strip()
        row = policy_rows[policy_id]
        column = columns[_INFORCE_FIELDS[field]]
        worksheet.cell(row, column).value = _coerce_inforce_value(field, edit.get("value"), [])


def _apply_parameter_changes(workbook: Any, input_path: Path, edits: dict[str, Any]) -> None:
    if not edits:
        return
    worksheet = workbook["Parameters"]
    for edit in edits.get("projection_settings", []):
        parameter = str(edit["parameter"]).strip()
        row = _find_marker_row(worksheet, parameter)
        if row is not None:
            worksheet.cell(row, 3).value = _coerce_parameter_scalar(
                parameter, edit.get("value"), []
            )

    _apply_gender_table_edits(worksheet, "Mortality Rates", edits.get("mortality_rates", []))
    _apply_gender_table_edits(worksheet, "Projection Scale", edits.get("projection_scale", []))
    _apply_random_number_edits(workbook, input_path, edits.get("random_numbers", []))


def _apply_gender_table_edits(worksheet: Any, marker: str, edits: list[dict[str, Any]]) -> None:
    start_row = _find_marker_row(worksheet, marker)
    if start_row is None:
        return
    for edit in edits:
        age = int(edit["age"])
        row = start_row + 1 + age
        if "male" in edit:
            worksheet.cell(row, 4).value = float(edit["male"])
        if "female" in edit:
            worksheet.cell(row, 5).value = float(edit["female"])


def _apply_random_number_edits(
    workbook: Any, input_path: Path, edits: list[dict[str, Any]]
) -> None:
    if not edits:
        return
    worksheet = workbook["Parameters"]
    data_workbook = openpyxl.load_workbook(
        input_path, read_only=True, data_only=True, keep_vba=True
    )
    try:
        start_row = _find_marker_row(data_workbook["Parameters"], "Random Numbers")
    finally:
        data_workbook.close()
    if start_row is None:
        return
    for edit in edits:
        row = start_row + int(edit["scenario"])
        column = 3 + int(edit["policy"])
        worksheet.cell(row, column).value = float(edit["value"])


def _apply_reporting_changes(workbook: Any, edits: list[dict[str, Any]]) -> None:
    if not edits:
        return
    worksheet = workbook["Reporting"]
    for edit in edits:
        section = str(edit["section"]).strip()
        field = str(edit["field"]).strip()
        row = _find_reporting_row(worksheet, section, field)
        if row is not None:
            worksheet.cell(row, 3).value = _coerce_reporting_value(
                section, field, edit.get("value"), []
            )


def _find_marker_row(worksheet: Any, marker: str) -> int | None:
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row, 1).value == marker:
            return row
    return None


def _find_reporting_row(worksheet: Any, section: str, field: str) -> int | None:
    section_row = _find_marker_row(worksheet, section)
    if section_row is None:
        return None
    if field == "Create?":
        return section_row
    for row in range(section_row + 1, worksheet.max_row + 1):
        first_cell = worksheet.cell(row, 1).value
        if first_cell in _REPORTING_FIELDS:
            return None
        if worksheet.cell(row, 2).value == field:
            return row
    return None


def _header_columns(worksheet: Any, header_row: int) -> dict[str, int]:
    return {
        str(worksheet.cell(header_row, column).value).strip(): column
        for column in range(1, worksheet.max_column + 1)
        if worksheet.cell(header_row, column).value is not None
    }


def _inforce_policy_rows(worksheet: Any) -> dict[int, int]:
    columns = _header_columns(worksheet, 1)
    policy_column = columns.get("Policy #", 1)
    policy_rows: dict[int, int] = {}
    for row in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row, policy_column).value
        if value is None or value == "":
            continue
        try:
            policy_rows[int(value)] = row
        except (TypeError, ValueError):
            continue
    return policy_rows


def _reject_formula(cell: Any, label: str, errors: list[str]) -> None:
    if isinstance(cell.value, str) and cell.value.startswith("="):
        errors.append(f"{label} is a formula cell and cannot be edited.")


def _int_value(value: Any, label: str, errors: list[str]) -> int | None:
    try:
        number = int(value)
    except (TypeError, ValueError):
        errors.append(f"{label} must be an integer.")
        return None
    return number


def _float_value(value: Any, label: str, errors: list[str]) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        errors.append(f"{label} must be numeric.")
        return None
    return number


def _rate_value(value: Any, label: str, errors: list[str]) -> float | None:
    number = _float_value(value, label, errors)
    if number is None:
        return None
    if not 0 <= number <= 1:
        errors.append(f"{label} must be between 0 and 1.")
        return None
    return number


def _yes_no_value(value: Any, label: str, errors: list[str]) -> str | None:
    text = str(value).strip()
    if text.lower() not in {"yes", "no"}:
        errors.append(f"{label} must be Yes or No.")
        return None
    return "Yes" if text.lower() == "yes" else "No"


def _coerce_inforce_value(field: str, value: Any, errors: list[str]) -> Any:
    if field == "YOB":
        year = _int_value(value, "YOB", errors)
        if year is not None and year <= 1800:
            errors.append("YOB must be greater than 1800.")
            return None
        return year
    if field == "Gender":
        gender = str(value).strip().upper()
        if gender not in {"M", "F"}:
            errors.append("Gender must be M or F.")
            return None
        return gender
    benefit = _float_value(value, "Annual Benefit", errors)
    if benefit is not None and benefit <= 0:
        errors.append("Annual Benefit must be greater than 0.")
        return None
    return benefit


def _coerce_parameter_scalar(parameter: str, value: Any, errors: list[str]) -> Any:
    if parameter == "Valuation Date":
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        try:
            return datetime.fromisoformat(str(value))
        except ValueError:
            errors.append("Valuation Date must be a valid date.")
            return None
    if parameter == "Last Projection Year":
        return _int_value(value, "Last Projection Year", errors)
    if parameter == "Which Random Numbers?":
        text = str(value).strip()
        if text.lower() not in {"seed", "table"}:
            errors.append("Which Random Numbers? must be Seed or Table.")
            return None
        return "Seed" if text.lower() == "seed" else "Table"
    return _int_value(value, "Random Numbers Seed", errors)


def _coerce_reporting_value(section: str, field: str, value: Any, errors: list[str]) -> Any:
    label = f"{section} {field}"
    if field in {"Create?", "Create graph?"}:
        return _yes_no_value(value, label, errors)
    if field == "Discount Rate":
        rate = _rate_value(value, label, errors)
        return rate
    return _int_value(value, label, errors)
