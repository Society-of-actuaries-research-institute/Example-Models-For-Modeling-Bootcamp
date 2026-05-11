"""Preview parsers for the desktop UI.

The UI only needs bounded table previews, so this module turns input and output
workbooks into small JSON-serialisable payloads. The model's typed loaders remain
the source of truth for workbook interpretation.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from mbc_model.data.loader import ExcelLoader
from mbc_model.data.models import ModelResults
from mbc_model.reporting.charts import (
    GRAPH_NOT_REQUESTED_MESSAGE,
    figure_to_data_url,
    make_dashboard_cash_flow_chart,
    make_scenario_cash_flows_chart,
)

PREVIEW_LIMIT = 100


def build_input_preview(path: Path, limit: int = PREVIEW_LIMIT) -> dict[str, Any]:
    """Return a bounded, JSON-ready preview of an input workbook."""
    path = path.resolve()
    loader = ExcelLoader(path)
    policies = loader.load_inforce()
    parameters, mortality_table, random_number_table = loader.load_parameters()
    reporting = loader.load_reporting()

    projection_start = parameters.valuation_date.year + 1
    projection_end = parameters.last_projection_year
    random_source = "Workbook table" if random_number_table is not None else "Seed"
    scenario_count = (
        int(random_number_table.shape[0])
        if random_number_table is not None
        else int(reporting.dashboard_scenarios)
    )

    inforce_rows = [
        [
            policy.policy_id,
            policy.yob,
            policy.gender,
            float(policy.annual_benefit),
        ]
        for policy in policies[:limit]
    ]

    mortality_rows = [
        [age, float(male), float(female)]
        for age, (male, female) in enumerate(
            zip(
                mortality_table.male_mortality_rates,
                mortality_table.female_mortality_rates,
            )
        )
    ]
    projection_scale_rows = [
        [age, float(male), float(female)]
        for age, (male, female) in enumerate(
            zip(
                mortality_table.male_projection_scale,
                mortality_table.female_projection_scale,
            )
        )
    ]

    if random_number_table is None:
        random_number_preview = _table_payload(
            ["Setting", "Value"],
            [
                ["Which Random Numbers?", "Seed"],
                ["Random Numbers Seed", parameters.random_seed],
            ],
            total_rows=2,
            total_columns=2,
            limit=limit,
        )
    else:
        random_rows: list[list[Any]] = []
        shown_scenarios = min(random_number_table.shape[0], limit)
        shown_policies = min(random_number_table.shape[1], limit - 1)
        for scenario_index in range(shown_scenarios):
            random_rows.append(
                [scenario_index + 1]
                + [
                    float(random_number_table[scenario_index, policy_index])
                    for policy_index in range(shown_policies)
                ]
            )
        random_number_preview = _table_payload(
            ["Scenario \\ Policy"] + [str(index + 1) for index in range(shown_policies)],
            random_rows,
            total_rows=int(random_number_table.shape[0]),
            total_columns=int(random_number_table.shape[1]) + 1,
            limit=limit,
        )

    reporting_sections = [
        {
            "title": "Policy Results",
            "headers": ["Parameter", "Value"],
            "rows": [
                ["Create?", _yes_no(reporting.create_policy_results)],
                ["Policy", reporting.policy_id],
                ["Scenario", reporting.policy_scenario_id],
            ],
        },
        {
            "title": "Scenario Results",
            "headers": ["Parameter", "Value"],
            "rows": [
                ["Create?", _yes_no(reporting.create_scenario_results)],
                ["Scenario", reporting.scenario_id],
                ["Create graph?", _yes_no(reporting.create_scenario_graph)],
            ],
        },
        {
            "title": "Total Results",
            "headers": ["Parameter", "Value"],
            "rows": [["Create?", _yes_no(reporting.create_total_results)]],
        },
        {
            "title": "Dashboard Results",
            "headers": ["Parameter", "Value"],
            "rows": [
                ["Create?", _yes_no(reporting.create_dashboard_results)],
                ["Scenarios", reporting.dashboard_scenarios],
                ["Policies", reporting.dashboard_policies],
                ["Discount Rate", reporting.discount_rate],
                ["Create graph?", _yes_no(reporting.create_dashboard_graph)],
            ],
        },
    ]

    return {
        "path": str(path),
        "file_name": path.name,
        "policy_count": len(policies),
        "scenario_count": scenario_count,
        "projection_start": projection_start,
        "projection_end": projection_end,
        "projection_years": projection_end - projection_start + 1,
        "discount_rate": float(reporting.discount_rate),
        "random_numbers": random_source,
        "inforce": _table_payload(
            ["Policy #", "YOB", "Gender", "Annual Benefit"],
            inforce_rows,
            total_rows=len(policies),
            total_columns=4,
            limit=limit,
        ),
        "parameters": {
            "projection_settings": _table_payload(
                ["Parameter", "Value"],
                [
                    ["Valuation Date", parameters.valuation_date.isoformat()],
                    ["Last Projection Year", parameters.last_projection_year],
                    [
                        "Which Random Numbers?",
                        "Table" if random_number_table is not None else "Seed",
                    ],
                    ["Random Numbers Seed", parameters.random_seed],
                ],
                total_rows=4,
                total_columns=2,
                limit=limit,
            ),
            "mortality_rates": _table_payload(
                ["Age \\ Gender", "Male", "Female"],
                mortality_rows,
                total_rows=len(mortality_rows),
                total_columns=3,
                limit=max(limit, len(mortality_rows)),
            ),
            "projection_scale": _table_payload(
                ["Age \\ Gender", "Male", "Female"],
                projection_scale_rows,
                total_rows=len(projection_scale_rows),
                total_columns=3,
                limit=max(limit, len(projection_scale_rows)),
            ),
            "random_numbers": random_number_preview,
        },
        "reporting": {"sections": reporting_sections},
    }


def build_output_preview(
    path: Path,
    results: ModelResults | None = None,
    limit: int = PREVIEW_LIMIT,
) -> dict[str, Any]:
    """Return a bounded, JSON-ready preview of a model output workbook."""
    path = path.resolve()
    workbook = openpyxl.load_workbook(path, data_only=True)
    dashboard = _parse_dashboard(workbook)
    total = _parse_total(workbook, limit)
    scenario = _parse_scenario(workbook, limit)
    policy = _parse_policy(workbook, limit)

    if results is not None:
        dashboard["chart"] = _dashboard_chart_payload(results)
        scenario["chart"] = _scenario_chart_payload(results)
    else:
        dashboard["chart"] = _chart_unavailable("Run the model to display this graph.")
        scenario["chart"] = _chart_unavailable("Run the model to display this graph.")

    return {
        "path": str(path),
        "file_name": path.name,
        "dashboard": dashboard,
        "total": total,
        "scenario": scenario,
        "policy": policy,
    }


def _parse_dashboard(workbook: openpyxl.Workbook) -> dict[str, Any]:
    if "Dashboard Results" not in workbook.sheetnames:
        return {"available": False, "rows": []}

    sheet = workbook["Dashboard Results"]
    discount_rate = sheet.cell(3, 3).value
    metric_names = [sheet.cell(5, column).value for column in range(1, sheet.max_column + 1)]
    metric_values = [sheet.cell(6, column).value for column in range(1, sheet.max_column + 1)]

    rows = [["Discount rate", _format_percent(discount_rate)]]
    for name, value in zip(metric_names, metric_values):
        if name is None or str(name).strip().lower() == "runtime":
            continue
        rows.append([str(name), _format_money(value)])
    return {"available": True, "rows": rows}


def _parse_total(workbook: openpyxl.Workbook, limit: int) -> dict[str, Any]:
    if "Total Results" not in workbook.sheetnames:
        return {"available": False}

    sheet = workbook["Total Results"]
    header_row = _find_row(sheet, "Year \\ Scen")
    if header_row is None:
        header_row = _find_row(sheet, "Year")

    pv_values = [
        _format_money(sheet.cell(2, column).value)
        for column in range(2, sheet.max_column + 1)
        if sheet.cell(2, column).value is not None
    ]
    pv_headers = [f"Scenario {index + 1}" for index in range(len(pv_values))]
    cash_flow = (
        _read_table(sheet, header_row, limit, money_columns=True) if header_row else _empty_table()
    )

    return {
        "available": True,
        "discount_rate": _format_percent(sheet.cell(1, 2).value),
        "pv_cash_flow": _table_payload(
            pv_headers,
            [pv_values] if pv_values else [],
            total_rows=1 if pv_values else 0,
            total_columns=len(pv_values),
            limit=limit,
        ),
        "cash_flow": cash_flow,
    }


def _parse_scenario(workbook: openpyxl.Workbook, limit: int) -> dict[str, Any]:
    if "Scenario Results" not in workbook.sheetnames:
        return {"available": False}

    sheet = workbook["Scenario Results"]
    header_row = _find_row(sheet, "Year \\ Policy") or _find_row(sheet, "Year")
    return {
        "available": True,
        "scenario": sheet.cell(1, 2).value,
        "cash_flow": (
            _read_table(sheet, header_row, limit, money_columns=True)
            if header_row
            else _empty_table()
        ),
    }


def _parse_policy(workbook: openpyxl.Workbook, limit: int) -> dict[str, Any]:
    if "Policy Results" not in workbook.sheetnames:
        return {"available": False}

    sheet = workbook["Policy Results"]
    header_row = _find_row(sheet, "Year")
    inputs: list[list[Any]] = [
        ["Policy", sheet.cell(1, 2).value, "Scenario", sheet.cell(1, 5).value],
        [
            "Gender",
            sheet.cell(2, 2).value,
            "Random Number",
            _format_decimal(sheet.cell(2, 5).value),
        ],
        ["Benefit", _format_money(sheet.cell(3, 2).value), "", ""],
    ]

    return {
        "available": True,
        "inputs": _table_payload(
            ["Field", "Value", "Field", "Value"],
            inputs,
            total_rows=len(inputs),
            total_columns=4,
            limit=limit,
        ),
        "calculations": (
            _read_table(sheet, header_row, limit, policy_columns=True)
            if header_row
            else _empty_table()
        ),
    }


def _scenario_chart_payload(results: ModelResults) -> dict[str, Any]:
    title = f"Scenario {results.config.scenario_id} Cash Flows by Policy"
    if not results.config.create_scenario_graph:
        return _chart_unavailable(GRAPH_NOT_REQUESTED_MESSAGE, title)

    figure = make_scenario_cash_flows_chart(results)
    if figure is None:
        return _chart_unavailable("Graph data is not available.", title)
    return {"available": True, "title": title, "data_url": figure_to_data_url(figure)}


def _dashboard_chart_payload(results: ModelResults) -> dict[str, Any]:
    title = "Cash Flow Projection by Scenario"
    if not results.config.create_dashboard_graph:
        return _chart_unavailable(GRAPH_NOT_REQUESTED_MESSAGE, title)

    figure = make_dashboard_cash_flow_chart(results)
    if figure is None:
        return _chart_unavailable("Graph data is not available.", title)
    return {"available": True, "title": title, "data_url": figure_to_data_url(figure)}


def _chart_unavailable(message: str, title: str = "") -> dict[str, Any]:
    return {"available": False, "title": title, "message": message}


def _read_table(
    sheet: Any,
    header_row: int,
    limit: int,
    money_columns: bool = False,
    policy_columns: bool = False,
) -> dict[str, Any]:
    max_column = min(sheet.max_column, limit)
    headers = [
        _format_cell(sheet.cell(header_row, column).value) for column in range(1, max_column + 1)
    ]
    rows: list[list[Any]] = []
    total_rows = 0
    for row_index in range(header_row + 1, sheet.max_row + 1):
        raw: list[Any] = [
            sheet.cell(row_index, column).value for column in range(1, sheet.max_column + 1)
        ]
        if all(value is None for value in raw):
            continue
        total_rows += 1
        if len(rows) >= limit:
            continue
        formatted_row = []
        for column_index, value in enumerate(raw[:max_column]):
            if money_columns and column_index > 0:
                formatted_row.append(_format_money(value))
            elif policy_columns and headers[column_index] == "Total Cash Flow":
                formatted_row.append(_format_money(value))
            elif policy_columns and isinstance(value, float) and column_index > 1:
                formatted_row.append(_format_decimal(value))
            else:
                formatted_row.append(_format_cell(value))
        rows.append(formatted_row)

    return _table_payload(
        headers,
        rows,
        total_rows=total_rows,
        total_columns=sheet.max_column,
        limit=limit,
    )


def _table_payload(
    headers: list[Any],
    rows: list[list[Any]],
    total_rows: int,
    total_columns: int,
    limit: int,
) -> dict[str, Any]:
    shown_headers = [_format_cell(header) for header in headers[:limit]]
    shown_rows = [[_format_cell(cell) for cell in row[:limit]] for row in rows[:limit]]
    return {
        "headers": shown_headers,
        "rows": shown_rows,
        "total_rows": int(total_rows),
        "total_columns": int(total_columns),
        "shown_rows": len(shown_rows),
        "shown_columns": len(shown_headers),
        "has_more_rows": total_rows > len(shown_rows),
        "has_more_columns": total_columns > len(shown_headers),
    }


def _empty_table() -> dict[str, Any]:
    return _table_payload([], [], total_rows=0, total_columns=0, limit=PREVIEW_LIMIT)


def _find_row(sheet: Any, first_cell_value: str) -> int | None:
    for row_index in range(1, sheet.max_row + 1):
        value = sheet.cell(row_index, 1).value
        if value == first_cell_value:
            return row_index
    return None


def _format_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _format_money(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"${round(float(value)):,.0f}"
    except (TypeError, ValueError):
        return str(value)


def _format_percent(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):.2%}"
    except (TypeError, ValueError):
        return str(value)


def _format_decimal(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):.6f}"
    except (TypeError, ValueError):
        return str(value)


def _yes_no(value: bool) -> str:
    return "Yes" if value else "No"
