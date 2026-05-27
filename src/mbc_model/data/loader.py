"""Excel loader: reads the input workbook and returns typed Python objects.

Parameters sheet layout (state-machine parser):
    Each logical table starts when column A contains the table name.
    The following rows (col A = None) contain data until the next table marker.

    Markers recognised:
        'Valuation Date'        -> scalar date     (column C)
        'Last Projection Year'  -> scalar int      (column C)
        'Which Random Numbers?' -> scalar string   (column C: "Table" or "Seed")
        'Random Numbers Seed'   -> scalar int      (column C)
        'Mortality Rates'       -> matrix float64  (col C = age, D = Male, E = Female)
        'Projection Scale'      -> matrix float64  (col C = age, D = Male, E = Female)
        'Random Numbers'        -> matrix float64  (col C = scenario, D+ = policies)

    State machine:
        Start in IDLE state.
        When a marker is found in column A, switch to the matching state.
        In table states, accumulate data rows (where column A is empty).
        Switch states when the next marker is found.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd

from mbc_model.data.models import (
    ModelParameters,
    MortalityTable,
    PolicyRecord,
    ReportingConfig,
)

_MIN_YOB: int = 1800  # Year of birth must be greater than this value
_VALID_GENDERS: frozenset[str] = frozenset({"M", "F"})  # Allowed gender values
_KNOWN_REPORTS: frozenset[str] = frozenset(
    {"Policy Results", "Scenario Results", "Total Results", "Dashboard Results"}
)  # Report section names recognised in the Reporting sheet


class ExcelLoader:
    """Reads a validated Excel workbook and returns typed model objects.

    All reading is read-only; the workbook is never modified.

    Args:
        path: Path to the xlsx input file.
    """

    def __init__(self, path: Path) -> None:
        self._path: Path = path  # Path to the input workbook (never written to)

    def load_inforce(self) -> list[PolicyRecord]:
        """Read the Inforce sheet and return a validated list of policies.

        Each row in the Inforce sheet becomes one PolicyRecord. Rows are
        validated for required columns, non-null values, valid gender,
        reasonable year of birth, and positive annual benefit.

        Returns:
            Ordered list of PolicyRecord, one per inforce row.

        Raises:
            ValueError: On missing columns, null values, invalid gender,
                YOB <= 1800, or non-positive benefit. Error messages include
                the Excel row number to help locate the problem.
        """
        # pandas reads the Inforce sheet into a DataFrame (a table with named columns)
        inforce_dataframe: pd.DataFrame = pd.read_excel(
            self._path,
            sheet_name="Inforce",
            header=0,  # First row is the column header
            engine="openpyxl",
        )

        required_columns: set[str] = {"Policy #", "YOB", "Gender", "Annual Benefit"}
        missing_columns: set[str] = required_columns - set(inforce_dataframe.columns)
        if missing_columns:
            raise ValueError(f"Inforce sheet: missing columns {missing_columns}.")

        if inforce_dataframe.empty:
            raise ValueError("Inforce sheet: no data rows found.")

        # Check for any null (empty) values in the required columns
        rows_with_null_values = inforce_dataframe[list(required_columns)].isnull().any(axis=1)
        if rows_with_null_values.any():
            # idxmax() returns the index of the first True value (first row with a null)
            excel_row_number: int = (
                int(rows_with_null_values.idxmax()) + 2
            )  # +1 header, +1 zero-index
            raise ValueError(
                f"Inforce sheet row {excel_row_number}: null value in required column."
            )

        inforce_policies: list[PolicyRecord] = []
        for row_index, row in inforce_dataframe.iterrows():
            excel_row_number = int(row_index) + 2  # type: ignore[arg-type]

            gender_value: str = str(row["Gender"]).strip()
            if gender_value not in _VALID_GENDERS:
                raise ValueError(
                    f"Inforce sheet row {excel_row_number}: "
                    f"Gender must be 'M' or 'F', got {gender_value!r}."
                )

            year_of_birth: int = int(row["YOB"])
            if year_of_birth <= _MIN_YOB:
                raise ValueError(
                    f"Inforce sheet row {excel_row_number}: "
                    f"YOB must be > {_MIN_YOB}, got {year_of_birth}."
                )

            annual_benefit_value: float = float(row["Annual Benefit"])
            if annual_benefit_value <= 0.0:
                raise ValueError(
                    f"Inforce sheet row {excel_row_number}: "
                    f"Annual Benefit must be > 0, got {annual_benefit_value}."
                )

            inforce_policies.append(
                PolicyRecord(
                    policy_id=int(row["Policy #"]),
                    yob=year_of_birth,
                    gender=gender_value,  # type: ignore[arg-type]
                    annual_benefit=annual_benefit_value,
                )
            )
        return inforce_policies

    def load_parameters(  # pylint: disable=too-many-branches,too-many-statements,too-many-locals
        self,
    ) -> tuple[ModelParameters, MortalityTable, np.ndarray | None]:
        """Read the Parameters sheet using a state-machine row parser.

        The parser scans column A for table name markers and accumulates data
        rows until the next marker is found.

        Returns:
            A tuple of:
                - ModelParameters: scalar settings (valuation date, last year, seed)
                - MortalityTable: mortality rates and projection scales by age
                - random_number_table: float64 array shape (n_scenarios, n_policies)
                    if "Which Random Numbers?" = "Table"; None if "Seed".

        Raises:
            KeyError: If a required table or scalar is missing from the sheet.
            ValueError: If "Which Random Numbers?" has an unrecognised value.
        """
        # openpyxl opens the workbook in read-only, data-only mode for speed.
        # data_only=True returns cell values instead of formula strings.
        workbook = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        parameters_worksheet = workbook["Parameters"]

        # Accumulators for each table type found in the sheet
        scalar_parameters: dict[str, Any] = {}
        mortality_rate_rows: list[tuple[float, float]] = []
        improvement_scale_rows: list[tuple[float, float]] = []
        random_number_rows: list[list[float]] = []

        # State machine constants: the parser is always in exactly one of these states
        STATE_IDLE: str = "idle"
        STATE_MORTALITY: str = "mortality"
        STATE_PROJECTION_SCALE: str = "projection_scale"
        STATE_RANDOM_NUMBERS: str = "random_numbers"

        parser_state: str = STATE_IDLE  # Start by scanning for the first marker

        for raw_spreadsheet_row in parameters_worksheet.iter_rows(values_only=True):
            row: list[Any] = list(raw_spreadsheet_row)

            # Column A contains either a table name marker or None (data row)
            table_name_column: Any = row[0] if row else None
            # Column C contains scalar values or the age index for table data
            value_column: Any = row[2] if len(row) > 2 else None

            # --- Detect table name markers in column A ----------------------------
            if table_name_column is not None and table_name_column != "Table":
                if table_name_column == "Valuation Date":
                    # Column C holds the valuation date value
                    valuation_date_value = value_column
                    if isinstance(valuation_date_value, datetime):
                        # openpyxl may return a datetime; convert to date only
                        scalar_parameters["valuation_date"] = valuation_date_value.date()
                    elif isinstance(valuation_date_value, date):
                        scalar_parameters["valuation_date"] = valuation_date_value
                    parser_state = STATE_IDLE

                elif table_name_column == "Last Projection Year":
                    scalar_parameters["last_projection_year"] = int(value_column)  # type: ignore[arg-type]
                    parser_state = STATE_IDLE

                elif table_name_column == "Which Random Numbers?":
                    # "Table" = use the Random Numbers table below
                    # "Seed"  = use the Random Numbers Seed scalar instead
                    scalar_parameters["which_random_numbers"] = (
                        str(value_column).strip() if value_column is not None else "Seed"
                    )
                    parser_state = STATE_IDLE

                elif table_name_column == "Random Numbers Seed":
                    scalar_parameters["random_seed"] = int(value_column)  # type: ignore[arg-type]
                    parser_state = STATE_IDLE

                elif table_name_column == "Mortality Rates":
                    parser_state = STATE_MORTALITY  # Next data rows are mortality rates

                elif table_name_column == "Projection Scale":
                    parser_state = STATE_PROJECTION_SCALE  # Next data rows are improvement scales

                elif table_name_column == "Random Numbers":
                    parser_state = STATE_RANDOM_NUMBERS  # Next data rows are random numbers

                continue  # Move to the next row after processing a marker

            # --- Skip unexpected non-None values in column A ----------------------
            if table_name_column is not None:
                continue

            # --- Accumulate data rows based on the current parser state -----------
            if parser_state == STATE_MORTALITY:
                # Column D = male qx value, column E = female qx value
                male_value: Any = row[3] if len(row) > 3 else None
                female_value: Any = row[4] if len(row) > 4 else None
                # Column C holds the age index (integer or float); skip non-numeric rows
                if isinstance(value_column, (int, float)) and male_value is not None:
                    mortality_rate_rows.append(
                        (
                            float(male_value),
                            float(female_value) if female_value is not None else 0.0,
                        )
                    )

            elif parser_state == STATE_PROJECTION_SCALE:
                # Column D = male improvement scale, column E = female improvement scale
                male_value = row[3] if len(row) > 3 else None
                female_value = row[4] if len(row) > 4 else None
                if isinstance(value_column, (int, float)) and male_value is not None:
                    improvement_scale_rows.append(
                        (
                            float(male_value),
                            float(female_value) if female_value is not None else 0.0,
                        )
                    )

            elif parser_state == STATE_RANDOM_NUMBERS:
                # Column C holds the scenario index; columns D onwards hold per-policy values
                policy_random_numbers: list[float] = [
                    float(cell_value) for cell_value in row[3:] if cell_value is not None
                ]
                if policy_random_numbers:
                    random_number_rows.append(policy_random_numbers)

        # --- Validate that all required data was found ----------------------------
        for required_key in ("valuation_date", "last_projection_year", "random_seed"):
            if required_key not in scalar_parameters:
                raise KeyError(
                    f"Parameters sheet: required scalar '{required_key}' not found. "
                    "Check that the table marker is present in column A."
                )
        if not mortality_rate_rows:
            raise KeyError("Parameters sheet: 'Mortality Rates' table not found or empty.")
        if not improvement_scale_rows:
            raise KeyError("Parameters sheet: 'Projection Scale' table not found or empty.")

        # Build the typed output objects from the accumulated data
        model_parameters: ModelParameters = ModelParameters(
            valuation_date=scalar_parameters["valuation_date"],
            last_projection_year=int(scalar_parameters["last_projection_year"]),
            random_seed=int(scalar_parameters["random_seed"]),
        )
        # numpy arrays allow vectorised mortality calculations in the engine
        mortality_table: MortalityTable = MortalityTable(
            male_mortality_rates=np.array(
                [row[0] for row in mortality_rate_rows], dtype=np.float64
            ),
            female_mortality_rates=np.array(
                [row[1] for row in mortality_rate_rows], dtype=np.float64
            ),
            male_projection_scale=np.array(
                [row[0] for row in improvement_scale_rows], dtype=np.float64
            ),
            female_projection_scale=np.array(
                [row[1] for row in improvement_scale_rows], dtype=np.float64
            ),
        )

        # Decide whether to return the random number table or None (seed mode)
        random_number_source: str = scalar_parameters.get("which_random_numbers", "Seed")
        if random_number_source.lower() == "table":
            # Convert the list of lists into a 2D NumPy array: shape (n_scenarios, n_policies)
            random_number_table: np.ndarray | None = (
                np.array(random_number_rows, dtype=np.float64) if random_number_rows else None
            )
        elif random_number_source.lower() == "seed":
            random_number_table = None  # Signal to the runner to use seeded generation
        else:
            raise ValueError(
                f"Parameters sheet: 'Which Random Numbers?' must be 'Table' or 'Seed', "
                f"got {random_number_source!r}."
            )

        return model_parameters, mortality_table, random_number_table

    def load_reporting(self) -> ReportingConfig:
        """Read the Reporting sheet and return the output configuration.

        The sheet is organised into named sections (Policy Results, Scenario Results,
        etc.). Each section has a "Create?" field and additional parameters.
        Missing boolean fields default to False; missing required numeric fields
        raise ValueError.

        Returns:
            ReportingConfig with all fields populated.

        Raises:
            ValueError: If a required numeric field (e.g. Discount Rate) is absent.
        """
        workbook = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        reporting_worksheet = workbook["Reporting"]

        # Build a nested dictionary: {section_name: {field_name: value}}
        report_section_data: dict[str, dict[str, Any]] = {}
        current_section_name: str | None = None

        for raw_spreadsheet_row in reporting_worksheet.iter_rows(values_only=True):
            row: list[Any] = list(raw_spreadsheet_row)
            section_name: Any = row[0] if row else None
            field_name: Any = row[1] if len(row) > 1 else None
            field_value: Any = row[2] if len(row) > 2 else None

            if section_name in _KNOWN_REPORTS:
                # Start of a new report section
                current_section_name = section_name
                report_section_data[current_section_name] = {}
                if field_name == "Create?":
                    # The "Create?" value appears on the same row as the section name
                    report_section_data[current_section_name]["Create?"] = field_value
                continue

            # Data rows within a section: column A is empty, column B is the field name
            if current_section_name and section_name is None and field_name is not None:
                report_section_data[current_section_name][str(field_name)] = field_value

        # Helper functions to extract values with type conversion and defaults

        def _bool(section: str, key: str, default: bool = False) -> bool:
            """Return True if the field value is the string 'yes' (case-insensitive)."""
            value = report_section_data.get(section, {}).get(key)
            if value is None:
                return default
            # Excel "Yes"/"No" drop-downs are read as strings; compare case-insensitively
            return str(value).strip().lower() == "yes"

        def _int(section: str, key: str, default: int | None = None) -> int:
            """Return the field value as an integer, or use the default if missing."""
            value = report_section_data.get(section, {}).get(key)
            if value is None:
                if default is not None:
                    return default
                raise ValueError(f"Reporting sheet: '{section}' is missing required field '{key}'.")
            return int(value)  # type: ignore[arg-type]

        def _float_required(section: str, key: str) -> float:
            """Return the field value as a float; raise ValueError if missing."""
            value = report_section_data.get(section, {}).get(key)
            if value is None:
                raise ValueError(f"Reporting sheet: '{section}' is missing required field '{key}'.")
            return float(value)  # type: ignore[arg-type]

        return ReportingConfig(
            create_policy_results=_bool("Policy Results", "Create?"),
            policy_id=_int("Policy Results", "Policy", default=1),
            policy_scenario_id=_int("Policy Results", "Scenario", default=1),
            create_scenario_results=_bool("Scenario Results", "Create?"),
            scenario_id=_int("Scenario Results", "Scenario", default=1),
            create_scenario_graph=_bool("Scenario Results", "Create graph?"),
            create_total_results=_bool("Total Results", "Create?"),
            create_dashboard_results=_bool("Dashboard Results", "Create?"),
            dashboard_scenarios=_int("Dashboard Results", "Scenarios", default=1),
            dashboard_policies=_int("Dashboard Results", "Policies", default=1),
            discount_rate=_float_required("Dashboard Results", "Discount Rate"),
            create_dashboard_graph=_bool("Dashboard Results", "Create graph?"),
        )
