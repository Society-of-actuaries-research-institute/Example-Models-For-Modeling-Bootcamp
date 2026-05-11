"""Excel report writer.

Writes a new ``outputs/results_{timestamp}.xlsx`` workbook with up to four
sheets as directed by ReportingConfig:

* **Policy Results**    - per-year detail for one (policy_id, scenario_id) pair.
* **Scenario Results**  - annual cash flows for every policy in one scenario.
* **Total Results**     - annual cash flows for every scenario.
* **Dashboard Results** - descriptive stats across scenarios, optionally with charts.
"""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

import numpy as np
import openpyxl
from openpyxl.drawing.image import Image as XlImage

from mbc_model.data.models import ModelResults
from mbc_model.reporting.charts import (
    figure_to_png_bytes,
    make_dashboard_cash_flow_chart,
    make_scenario_cash_flows_chart,
)

if TYPE_CHECKING:
    from matplotlib.figure import Figure


class ReportWriter:
    """Writes model results to a new Excel workbook."""

    def __init__(self, output_dir: Path = Path("outputs")) -> None:
        self._output_dir: Path = output_dir

    def write(self, results: ModelResults) -> Path:
        """Write all requested sheets and return the path to the new workbook."""
        self._output_dir.mkdir(parents=True, exist_ok=True)
        timestamp_string: str = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path: Path = self._output_dir / f"results_{timestamp_string}.xlsx"

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # type: ignore[arg-type]

        reporting_config = results.config
        if reporting_config.create_policy_results and results.policy_detail is not None:
            self._write_policy_results(workbook, results)
        if reporting_config.create_scenario_results:
            self._write_scenario_results(workbook, results)
        if reporting_config.create_total_results:
            self._write_total_results(workbook, results)
        if reporting_config.create_dashboard_results:
            self._write_dashboard_results(workbook, results)

        if not workbook.sheetnames:
            raise ValueError(
                "ReportingConfig has no output sheets enabled. "
                "Set at least one 'Create?' field to 'Yes' in the Reporting sheet."
            )

        workbook.save(output_path)
        return output_path.resolve()

    def _write_policy_results(self, workbook: openpyxl.Workbook, results: ModelResults) -> None:
        """Write per-year detail for one policy/scenario pair."""
        policy_detail = results.policy_detail
        assert policy_detail is not None
        matched_policy = next(p for p in results.policies if p.policy_id == policy_detail.policy_id)

        worksheet = workbook.create_sheet("Policy Results")
        worksheet.append(
            ["Policy", policy_detail.policy_id, None, "Scenario", policy_detail.scenario_id]
        )
        worksheet.append(
            [
                "Gender",
                matched_policy.gender,
                None,
                "Random Number",
                policy_detail.random_number,
            ]
        )
        worksheet.append(
            [
                "Benefit",
                matched_policy.annual_benefit,
                None,
                None,
                None,
                None,
                "Cumulative Probability of",
            ]
        )
        worksheet.append([None, None, None, None, None, None, "Survival", "Death", "Survive = 1"])
        worksheet.append(
            [
                "Year",
                "Age",
                "Base Qx",
                "Improvement",
                "Improved Qx",
                "Px",
                "tPx",
                "1 - tPx",
                "Dead = 0",
                "Total Cash Flow",
            ]
        )

        for year_index, projection_year in enumerate(policy_detail.projection_years):
            worksheet.append(
                [
                    int(projection_year),
                    int(policy_detail.attained_ages[year_index]),
                    float(policy_detail.base_qx[year_index]),
                    float(policy_detail.improvement_factor[year_index]),
                    float(policy_detail.improved_qx[year_index]),
                    float(policy_detail.px[year_index]),
                    float(policy_detail.cumulative_probability_of_survival_tPx[year_index]),
                    float(policy_detail.cumulative_probability_of_death_1_minus_tPx[year_index]),
                    int(policy_detail.survive_1_dead_0[year_index]),
                    float(policy_detail.total_cash_flow[year_index]),
                ]
            )

    def _write_scenario_results(self, workbook: openpyxl.Workbook, results: ModelResults) -> None:
        """Write cash flows by policy for one selected scenario."""
        reporting_config = results.config
        scenario_index: int = reporting_config.scenario_id - 1
        inforce_policies = results.policies
        projection_years = results.projection_years
        policy_cash_flows = results.scenario_policy_cash_flows
        total_scenario_cash_flow = results.scenario_cash_flows[scenario_index]
        show_total_only: bool = len(inforce_policies) > 10

        worksheet = workbook.create_sheet("Scenario Results")
        worksheet.append(["Scenario", reporting_config.scenario_id])
        worksheet.append([None, "Cash Flow Projection by Policy"])
        if show_total_only:
            worksheet.append(["Year", "Total"])
        else:
            worksheet.append(
                ["Year \\ Policy"] + [p.policy_id for p in inforce_policies] + ["Total"]
            )

        for year_index, projection_year in enumerate(projection_years):
            if show_total_only:
                worksheet.append(
                    [
                        int(projection_year),
                        float(total_scenario_cash_flow[year_index]),
                    ]
                )
            else:
                if policy_cash_flows is not None:
                    cash_flow_per_policy: list[float] = [
                        float(policy_cash_flows[policy_index, year_index])
                        for policy_index in range(len(inforce_policies))
                    ]
                else:
                    cash_flow_per_policy = [""] * len(inforce_policies)  # type: ignore[list-item]
                worksheet.append(
                    [int(projection_year)]
                    + cash_flow_per_policy
                    + [float(total_scenario_cash_flow[year_index])]
                )

        if reporting_config.create_scenario_graph:
            figure = make_scenario_cash_flows_chart(results)
            if figure is not None:
                self._embed_chart(worksheet, figure, anchor="A" + str(worksheet.max_row + 2))

    def _write_total_results(self, workbook: openpyxl.Workbook, results: ModelResults) -> None:
        """Write total cash flows by scenario for each year."""
        worksheet = workbook.create_sheet("Total Results")
        number_of_scenarios: int = results.scenario_cash_flows.shape[0]
        projection_years = results.projection_years
        discount_rate: float = results.config.discount_rate

        valuation_year: int = int(projection_years[0]) - 1
        time_periods_from_valuation: np.ndarray = (projection_years - valuation_year).astype(
            np.float64
        )
        discount_factors_by_year: np.ndarray = (
            1.0 / (1.0 + discount_rate) ** time_periods_from_valuation
        )
        present_values_by_scenario: np.ndarray = (
            results.scenario_cash_flows @ discount_factors_by_year
        )

        worksheet.append(["Discount Rate", discount_rate])
        worksheet.append(
            [
                "PV Cash Flow",
                *[
                    float(present_values_by_scenario[scenario_index])
                    for scenario_index in range(number_of_scenarios)
                ],
            ]
        )

        if number_of_scenarios <= 25:
            worksheet.append([None, "Total Cash Flow by Scenario"])
            worksheet.append(["Year \\ Scen"] + list(range(1, number_of_scenarios + 1)))
            for year_index, projection_year in enumerate(projection_years):
                data_row: list[float | int] = [int(projection_year)] + [
                    float(results.scenario_cash_flows[scenario_index, year_index])
                    for scenario_index in range(number_of_scenarios)
                ]
                worksheet.append(data_row)

    def _write_dashboard_results(self, workbook: openpyxl.Workbook, results: ModelResults) -> None:
        """Write summary statistics across scenarios."""
        reporting_config = results.config
        present_values_by_scenario = results.pv_by_scenario
        if present_values_by_scenario is None:
            return

        number_of_scenarios_used: int = min(
            reporting_config.dashboard_scenarios, len(present_values_by_scenario)
        )
        number_of_policies_used: int = min(
            reporting_config.dashboard_policies, len(results.policies)
        )
        present_values_subset: np.ndarray = present_values_by_scenario[:number_of_scenarios_used]

        total_runtime_seconds: int = int(results.runtime_seconds)
        runtime_formatted: str = (
            f"{total_runtime_seconds // 3600:02d}:"
            f"{(total_runtime_seconds % 3600) // 60:02d}:"
            f"{total_runtime_seconds % 60:02d}"
        )

        worksheet = workbook.create_sheet("Dashboard Results")
        worksheet.append(["Number of Scenarios", None, number_of_scenarios_used])
        worksheet.append(["Number of Policies", None, number_of_policies_used])
        worksheet.append(["Discount rate", None, reporting_config.discount_rate])
        worksheet.append(["PV Cash Flow Statistics"])
        worksheet.append(["Mean", "Median", "Std Dev", "Min", "Max", "Runtime"])
        worksheet.append(
            [
                float(np.mean(present_values_subset)),
                float(np.median(present_values_subset)),
                float(np.std(present_values_subset)),
                float(np.min(present_values_subset)),
                float(np.max(present_values_subset)),
                runtime_formatted,
            ]
        )

        if reporting_config.create_dashboard_graph:
            figure = make_dashboard_cash_flow_chart(results)
            if figure is not None:
                self._embed_chart(worksheet, figure, anchor="A" + str(worksheet.max_row + 2))

    @staticmethod
    def _embed_chart(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        figure: "Figure",
        anchor: str,
    ) -> None:
        """Save a matplotlib figure as a PNG and embed it in a worksheet."""
        chart_image = XlImage(io.BytesIO(figure_to_png_bytes(figure)))
        worksheet.add_image(chart_image, anchor)
