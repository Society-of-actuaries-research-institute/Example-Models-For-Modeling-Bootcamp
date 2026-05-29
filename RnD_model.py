# R&D Model Example for SOA Modeling Bootcamp 2026
# Written by Igor Nikitin, ASA, MAAA

# To run this script you will need:
#  - installation of Python 3.10 or higher (download from https://www.python.org/downloads/)
#  - installation of the required Python libraries

#  Once Python and the required libraries are installed, you can run this script from command prompt, vs code, or any other Python IDE.

#  To run from Command Prompt:
#  - navigate to the directory where this file is located (e.g., cd C:\Users\nikit\Desktop\Modeling Bootcamp)
#  - run the command: python RnD_model.py "inputs/Input 10 pol 25 scen table.xlsx"

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import io
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.drawing.image import Image as XlImage
from pathlib import Path
import sys
import time

# =========
# 1. Inputs
# =========

Valuation_Date = None
Last_Projection_Year = None

@dataclass(frozen=True)
class PolicyRecord:
    Policy: int
    YOB: int
    Gender: str
    Annual_Benefit: int

Inforce = []
Mortality_Rates = {"Male": [], "Female": []}
Projection_Scale = {"Male": [], "Female": []}
Random_Numbers = []
Which_Random_Numbers = "Table"
Random_Numbers_Seed = 1
Reporting_Config = {}


def load_input_workbook(input_path: Path) -> None:
    global Valuation_Date
    global Last_Projection_Year
    global Inforce
    global Mortality_Rates
    global Projection_Scale
    global Random_Numbers
    global Which_Random_Numbers
    global Random_Numbers_Seed
    global Reporting_Config

    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    inforce_sheet = workbook["Inforce"]
    Inforce = []
    for row in inforce_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        Inforce.append(
            PolicyRecord(
                Policy=int(row[0]),
                YOB=int(row[1]),
                Gender=str(row[2]),
                Annual_Benefit=int(row[3]),
            )
        )

    parameters_sheet = workbook["Parameters"]
    Mortality_Rates = {"Male": [], "Female": []}
    Projection_Scale = {"Male": [], "Female": []}
    Random_Numbers = []
    section = ""

    for row in parameters_sheet.iter_rows(values_only=True):
        table_name = row[0]
        value = row[2] if len(row) > 2 else None

        if table_name is not None and table_name != "Table":
            if table_name == "Valuation Date":
                Valuation_Date = value.date() if isinstance(value, datetime) else value
                section = ""
            elif table_name == "Last Projection Year":
                Last_Projection_Year = int(value)
                section = ""
            elif table_name == "Which Random Numbers?":
                Which_Random_Numbers = str(value)
                section = ""
            elif table_name == "Random Numbers Seed":
                Random_Numbers_Seed = int(value)
                section = ""
            elif table_name == "Mortality Rates":
                section = "Mortality Rates"
            elif table_name == "Projection Scale":
                section = "Projection Scale"
            elif table_name == "Random Numbers":
                section = "Random Numbers"
            continue

        if table_name is not None:
            continue

        if section == "Mortality Rates" and isinstance(value, (int, float)):
            Mortality_Rates["Male"].append(float(row[3]))
            Mortality_Rates["Female"].append(float(row[4]))
        elif section == "Projection Scale" and isinstance(value, (int, float)):
            Projection_Scale["Male"].append(float(row[3]))
            Projection_Scale["Female"].append(float(row[4]))
        elif section == "Random Numbers":
            random_numbers_for_scenario = [float(cell) for cell in row[3:] if cell is not None]
            if random_numbers_for_scenario:
                Random_Numbers.append(random_numbers_for_scenario)

    reporting_sheet = workbook["Reporting"]
    Reporting_Config = {}
    current_report = ""
    known_reports = {"Policy Results", "Scenario Results", "Total Results", "Dashboard Results"}

    for row in reporting_sheet.iter_rows(values_only=True):
        report_name = row[0]
        parameter_name = row[1] if len(row) > 1 else None
        parameter_value = row[2] if len(row) > 2 else None

        if report_name in known_reports:
            current_report = str(report_name)
            Reporting_Config[current_report] = {}
            if parameter_name == "Create?":
                Reporting_Config[current_report]["Create?"] = parameter_value
            continue

        if current_report and report_name is None and parameter_name is not None:
            Reporting_Config[current_report][str(parameter_name)] = parameter_value

def get_random_number(policy_number: int, scenario_number: int) -> float:
    if Which_Random_Numbers.strip().lower() == "seed":
        random_number_generator = np.random.default_rng(Random_Numbers_Seed + policy_number - 1)
        return float(random_number_generator.random(scenario_number)[-1])
    return float(Random_Numbers[scenario_number - 1][policy_number - 1])


def report_value(report_name: str, parameter_name: str, default):
    return Reporting_Config.get(report_name, {}).get(parameter_name, default)


def should_create_report(report_name: str) -> bool:
    return str(report_value(report_name, "Create?", "No")).strip().lower() == "yes"


def number_of_scenarios_to_run() -> int:
    if Which_Random_Numbers.strip().lower() == "table":
        return len(Random_Numbers)
    return int(report_value("Dashboard Results", "Scenarios", 1))


# ==============================================================================
# 2. Policy calculations for a given Scenario (Calculations tab in the workbook)
# ==============================================================================

def calculate_policy(policy_number: int, scenario_number: int) -> list[dict[str, int]]:
    policy = Inforce[policy_number - 1]
    random_number = get_random_number(policy_number, scenario_number)
    if policy.Gender == "M":
        base_qx_table = Mortality_Rates["Male"]
        improvement_table = Projection_Scale["Male"]
    if policy.Gender == "F":
        base_qx_table = Mortality_Rates["Female"]
        improvement_table = Projection_Scale["Female"]
    policy_results = []
    cumulative_probability_of_survival_tPx = 1.0
    first_projection_year = Valuation_Date.year + 1
    projection_years = list(range(first_projection_year, Last_Projection_Year + 1))

    for projection_year in projection_years:
        age = projection_year - policy.YOB
        if age >= len(base_qx_table):
            base_qx = 1.0
            improvement = 1.0
        else:               
            base_qx = base_qx_table[age]
            improvement = (1- improvement_table[age]) ** (projection_year - 2012)
        improved_qx = base_qx * improvement
        px = 1 - improved_qx
        cumulative_probability_of_survival_tPx *= px
        cumulative_probability_of_death_1_minus_tPx = 1 - cumulative_probability_of_survival_tPx
        survive_1_dead_0 = int(random_number > cumulative_probability_of_death_1_minus_tPx)
        cash_flow = policy.Annual_Benefit * survive_1_dead_0

        policy_results.append({
            "Year": projection_year,
            "Age": age,
            "Base_Qx": base_qx,
            "Improvement": improvement,
            "Improved_Qx": improved_qx,
            "Px": px,
            "Cumulative_Probability_of_Survival_tPx": cumulative_probability_of_survival_tPx,
            "Cumulative_Probability_of_Death_1_minus_tPx": cumulative_probability_of_death_1_minus_tPx,
            "Survive_1_Dead_0": survive_1_dead_0,
            "Total_Cash_Flow": cash_flow,
        })

    return policy_results

# =============================================
# Print results for a given policy and scenario
# =============================================

def print_policy_results(policy_number: int, scenario_number: int) -> None:
    policy_results = calculate_policy(policy_number, scenario_number)
    headers = list(policy_results[0].keys())
    print("\nPolicy Results for Policy Number:", policy_number, "and Scenario Number:", scenario_number)
    print("\t".join(headers))
    for row in policy_results:
        print("\t".join(str(row[header]) for header in headers))

# ==============================================
# 3. Calculate Scenario Results for all policies
# ==============================================

def calculate_scenario_cash_flows(
    scenario_number: int, number_of_policies: int
) -> list[dict[str, int]]:
    cash_flows_by_policy = {
        policy_number: calculate_policy(policy_number, scenario_number)
        for policy_number in range(1, number_of_policies + 1)
    }

    scenario_total_cash_flows = []

    for projection_year in range(len(list(cash_flows_by_policy.values())[0])):
        total_cash_flow_in_a_given_year = sum(
            cash_flows_by_policy[policy_number][projection_year]["Total_Cash_Flow"]
            for policy_number in range(1, number_of_policies + 1)
        )
        scenario_total_cash_flows.append(total_cash_flow_in_a_given_year)

    scenario_results = []
    first_projection_year = Valuation_Date.year + 1
    for projection_year in range(len(scenario_total_cash_flows)):
        year = first_projection_year + projection_year
        cash_flows_for_year = {
            f"Policy_{policy_number}": cash_flows_by_policy[policy_number][projection_year]["Total_Cash_Flow"]
            for policy_number in range(1, number_of_policies + 1)
        }
        total_cash_flow_for_year = scenario_total_cash_flows[projection_year]
        scenario_results.append({"Year": year, **cash_flows_for_year, "Total": total_cash_flow_for_year})
    
    return scenario_results

# =======================================
# Print scenario results for all policies
# =======================================

def print_scenario_results(
    scenario_number: int, number_of_policies: int
) -> None:
    
    scenario_results = calculate_scenario_cash_flows(scenario_number, number_of_policies)

    headers = ["Year"] + [
        f"Policy_{policy_number}"
        for policy_number in range(1, number_of_policies + 1)
    ] + ["Total"]
    print("\nScenario Results for Scenario Number:", scenario_number)
    print("\t".join(headers))
    for row in scenario_results:
        print("\t".join(str(row[header]) for header in headers))

# ==========================================
# Plot scenario results for all policies
# ==========================================

def plot_scenario_results(scenario_number: int, number_of_policies: int) -> None:
    scenario_results = calculate_scenario_cash_flows(scenario_number, number_of_policies)
    years = [row["Year"] for row in scenario_results]
    policy_cash_flows = {
        f"Policy_{policy_number}": [row[f"Policy_{policy_number}"] for row in scenario_results]
        for policy_number in range(1, number_of_policies + 1)
    }
    total_cash_flows = [row["Total"] for row in scenario_results]

    bottom = [0] * len(years)
    for policy_number in range(1, number_of_policies + 1):
        policy_label = f"Policy_{policy_number}"
        plt.bar(years, policy_cash_flows[policy_label], bottom=bottom, label=policy_label)
        bottom = [bottom[i] + policy_cash_flows[policy_label][i] for i in range(len(bottom))]

    plt.plot(years, total_cash_flows, color="black", marker="o", linestyle="--", label="Total")
    plt.xlabel("Year")
    plt.ylabel("Cash Flow")
    plt.title(f"Scenario {scenario_number} Cash Flows by Policy")
    plt.legend()
    plt.show()

# ======================================
# 4. Calculate Total Results by Scenario
# ======================================

def calculate_total_results_by_scenario(number_of_scenarios: int, number_of_policies: int) -> list[dict[str, int]]:
    scenario_cash_flows = {
        scenario_number: calculate_scenario_cash_flows(scenario_number, number_of_policies)
        for scenario_number in range(1, number_of_scenarios + 1)
    }
    total_results_by_scenario = []
    first_projection_year = Valuation_Date.year + 1
    for projection_year in range(Last_Projection_Year - Valuation_Date.year):
        year = first_projection_year + projection_year
        total_cash_flows_for_year = {
            f"Scenario_{scenario_number}": scenario_cash_flows[scenario_number][projection_year]["Total"]
            for scenario_number in range(1, number_of_scenarios + 1)
        }
        total_results_by_scenario.append({"Year": year, **total_cash_flows_for_year})
    
    return total_results_by_scenario

# ==============================
# Print results across scenarios
# ==============================

def print_results_across_scenarios(number_of_scenarios: int, number_of_policies: int) -> None:
    total_results_by_scenario = calculate_total_results_by_scenario(number_of_scenarios, number_of_policies)

    headers = ["Year"] + [
        f"Scenario_{scenario_number}"
        for scenario_number in range(1, number_of_scenarios + 1)
    ]
    print("\nResults Across Scenarios:")
    print("\t".join(headers))
    for row in total_results_by_scenario:
        print("\t".join(str(row[header]) for header in headers))

# ===========================
# Plot cash flows by scenario
# ===========================
def plot_cash_flows_by_scenario(number_of_scenarios: int, number_of_policies: int) -> None:
    cash_flows_by_scenario = calculate_total_results_by_scenario(number_of_scenarios, number_of_policies)
    years = [row["Year"] for row in cash_flows_by_scenario]
    for scenario_number in range(1, number_of_scenarios + 1):
        pv_cash_flows = [row[f"Scenario_{scenario_number}"] for row in cash_flows_by_scenario]
        plt.plot(years, pv_cash_flows, label=f"Scenario_{scenario_number}")
    plt.xlabel("Year")
    plt.ylabel("PV Cash Flow")
    plt.title("Cash Flow Projection by Scenario")
    plt.legend()
    plt.show()

# ======================================
# 5. Calculate PV Cash Flows by Scenario
# ======================================

def calculate_pv_cash_flows_by_scenario(number_of_scenarios: int, number_of_policies: int, discount_rate: float) -> dict[int]:
    total_results_by_scenario = calculate_total_results_by_scenario(number_of_scenarios, number_of_policies)
    pv_cash_flows_by_scenario = {}
    valuation_year = Valuation_Date.year

    for scenario_number in range(1, number_of_scenarios + 1):
        cash_flows = [row[f"Scenario_{scenario_number}"] for row in total_results_by_scenario]
        pv_cash_flows = 0
        for i, cash_flow in enumerate(cash_flows):
            year = valuation_year + 1 + i
            pv_cash_flows += cash_flow / (1 + discount_rate) ** (year - valuation_year)
        pv_cash_flows_by_scenario[scenario_number] = pv_cash_flows
    return pv_cash_flows_by_scenario

# ===============================
# Print PV cash flows by scenario
# ===============================

def print_pv_cash_flows_by_scenario(number_of_scenarios: int, number_of_policies: int, discount_rate: float) -> None:
    pv_cash_flows_by_scenario = calculate_pv_cash_flows_by_scenario(
        number_of_scenarios, number_of_policies, discount_rate
    )
    print("\nPV of Cash Flows by Scenario:")
    print("Scenario\tPV_Cash_Flow")
    for scenario_number, pv_cash_flow in pv_cash_flows_by_scenario.items():
        print(f"Scenario_{scenario_number}\t{pv_cash_flow:.2f}")

# ================================================
# Calculate the dashboard results across scenarios
# ================================================

def calculate_dashboard_results_by_scenario(number_of_scenarios: int, number_of_policies: int, discount_rate: float) -> dict[str, float]:
    pv_cash_flows_by_scenario = calculate_pv_cash_flows_by_scenario(
        number_of_scenarios, number_of_policies, discount_rate
    )
    pv_cash_flow_values = list(pv_cash_flows_by_scenario.values())
    mean_pv_cash_flow = sum(pv_cash_flow_values) / len(pv_cash_flow_values)
    median_pv_cash_flow = sorted(pv_cash_flow_values)[len(pv_cash_flow_values) // 2]
    std_dev_pv_cash_flow = (sum((x - mean_pv_cash_flow) ** 2 for x in pv_cash_flow_values) / len(pv_cash_flow_values)) ** 0.5
    min_pv_cash_flow = min(pv_cash_flow_values)
    max_pv_cash_flow = max(pv_cash_flow_values)

    return {
        "Mean_PV_Cash_Flow": mean_pv_cash_flow,
        "Median_PV_Cash_Flow": median_pv_cash_flow,
        "Std_Dev_PV_Cash_Flow": std_dev_pv_cash_flow,
        "Min_PV_Cash_Flow": min_pv_cash_flow,
        "Max_PV_Cash_Flow": max_pv_cash_flow,
    }

#=========================================
# Print dashboard results across scenarios
#=========================================

def print_dashboard_results_by_scenario(number_of_scenarios: int, number_of_policies: int, discount_rate: float) -> None:
    dashboard_results = calculate_dashboard_results_by_scenario(
        number_of_scenarios, number_of_policies, discount_rate
    )
    print("\nDashboard Results:")
    for key, value in dashboard_results.items():
        print(f"{key}: {value:.2f}")

# =================================
# Generate output workbook as xlsx
# =================================

def generate_output_xlsx(start_time: float, output_dir: Path = Path("outputs")) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp_string = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"rnd_results_{timestamp_string}.xlsx"
    counter = 1
    while output_path.exists():
        output_path = output_dir / f"rnd_results_{timestamp_string}_{counter}.xlsx"
        counter += 1

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # type: ignore[arg-type]

    number_of_policies = len(Inforce)
    number_of_scenarios = number_of_scenarios_to_run()
    discount_rate = float(report_value("Dashboard Results", "Discount Rate", 0.04))
    total_results_by_scenario = None
    pv_cash_flows_by_scenario = None

    if should_create_report("Policy Results"):
        policy_number = int(report_value("Policy Results", "Policy", 1))
        scenario_number = int(report_value("Policy Results", "Scenario", 1))
        write_policy_results_sheet(workbook, policy_number, scenario_number)

    if should_create_report("Scenario Results"):
        scenario_number = int(report_value("Scenario Results", "Scenario", 1))
        scenario_results = calculate_scenario_cash_flows(scenario_number, number_of_policies)
        write_scenario_results_sheet(workbook, scenario_number, number_of_policies, scenario_results)

    if should_create_report("Total Results") or should_create_report("Dashboard Results"):
        total_results_by_scenario = calculate_total_results_by_scenario(
            number_of_scenarios, number_of_policies
        )
        pv_cash_flows_by_scenario = calculate_pv_cash_flows_by_scenario(
            number_of_scenarios, number_of_policies, discount_rate
        )

    if should_create_report("Total Results"):
        assert total_results_by_scenario is not None
        assert pv_cash_flows_by_scenario is not None
        write_total_results_sheet(
            workbook,
            number_of_scenarios,
            discount_rate,
            total_results_by_scenario,
            pv_cash_flows_by_scenario,
        )

    if should_create_report("Dashboard Results"):
        assert total_results_by_scenario is not None
        assert pv_cash_flows_by_scenario is not None
        dashboard_scenarios = int(report_value("Dashboard Results", "Scenarios", number_of_scenarios))
        dashboard_policies = int(report_value("Dashboard Results", "Policies", number_of_policies))
        runtime_seconds = time.time() - start_time
        write_dashboard_results_sheet(
            workbook,
            min(dashboard_scenarios, number_of_scenarios),
            min(dashboard_policies, number_of_policies),
            discount_rate,
            runtime_seconds,
            total_results_by_scenario,
            pv_cash_flows_by_scenario,
        )

    workbook.save(output_path)
    return output_path.resolve()


def write_policy_results_sheet(
    workbook: openpyxl.Workbook, policy_number: int, scenario_number: int
) -> None:
    policy = Inforce[policy_number - 1]
    random_number = get_random_number(policy_number, scenario_number)
    policy_results = calculate_policy(policy_number, scenario_number)

    worksheet = workbook.create_sheet("Policy Results")
    worksheet.append(["Policy", policy_number, None, "Scenario", scenario_number])
    worksheet.append(["Gender", policy.Gender, None, "Random Number", random_number])
    worksheet.append(
        ["Benefit", policy.Annual_Benefit, None, None, None, None, "Cumulative Probability of"]
    )
    worksheet.append([None, None, None, None, None, None, "Survival", "Death", "Survive = 1"])
    worksheet.append(
        [
            "Year",
            "Age",
            "Base Qx",
            "Improvement",
            "Improved Qx",
            "Px",
            "tPx",
            "1 - tPx",
            "Dead = 0",
            "Total Cash Flow",
        ]
    )

    for row in policy_results:
        worksheet.append(
            [
                row["Year"],
                row["Age"],
                row["Base_Qx"],
                row["Improvement"],
                row["Improved_Qx"],
                row["Px"],
                row["Cumulative_Probability_of_Survival_tPx"],
                row["Cumulative_Probability_of_Death_1_minus_tPx"],
                row["Survive_1_Dead_0"],
                row["Total_Cash_Flow"],
            ]
        )


def write_scenario_results_sheet(
    workbook: openpyxl.Workbook,
    scenario_number: int,
    number_of_policies: int,
    scenario_results: list[dict[str, int]],
) -> None:
    show_total_only = number_of_policies > 10

    worksheet = workbook.create_sheet("Scenario Results")
    worksheet.append(["Scenario", scenario_number])
    worksheet.append([None, "Cash Flow Projection by Policy"])
    if show_total_only:
        worksheet.append(["Year", "Total"])
    else:
        worksheet.append(["Year \\ Policy"] + [policy.Policy for policy in Inforce] + ["Total"])

    for row in scenario_results:
        if show_total_only:
            worksheet.append([row["Year"], row["Total"]])
        else:
            worksheet.append(
                [row["Year"]]
                + [row[f"Policy_{policy.Policy}"] for policy in Inforce]
                + [row["Total"]]
            )

    if str(report_value("Scenario Results", "Create graph?", "No")).strip().lower() == "yes":
        figure = make_scenario_results_figure(scenario_number, number_of_policies, scenario_results)
        if figure is not None:
            embed_chart(worksheet, figure, "A" + str(worksheet.max_row + 2))


def write_total_results_sheet(
    workbook: openpyxl.Workbook,
    number_of_scenarios: int,
    discount_rate: float,
    total_results_by_scenario: list[dict[str, int]],
    pv_cash_flows_by_scenario: dict[int, float],
) -> None:
    worksheet = workbook.create_sheet("Total Results")
    worksheet.append(["Discount Rate", discount_rate])
    worksheet.append(
        ["PV Cash Flow"]
        + [
            pv_cash_flows_by_scenario[scenario_number]
            for scenario_number in range(1, number_of_scenarios + 1)
        ]
    )

    if number_of_scenarios <= 25:
        worksheet.append([None, "Total Cash Flow by Scenario"])
        worksheet.append(["Year \\ Scen"] + list(range(1, number_of_scenarios + 1)))
        for row in total_results_by_scenario:
            worksheet.append(
                [row["Year"]]
                + [
                    row[f"Scenario_{scenario_number}"]
                    for scenario_number in range(1, number_of_scenarios + 1)
                ]
            )


def write_dashboard_results_sheet(
    workbook: openpyxl.Workbook,
    number_of_scenarios: int,
    number_of_policies: int,
    discount_rate: float,
    runtime_seconds: float,
    total_results_by_scenario: list[dict[str, int]],
    pv_cash_flows_by_scenario: dict[int, float] | None = None,
) -> None:
    assert pv_cash_flows_by_scenario is not None
    pv_values = [pv_cash_flows_by_scenario[i] for i in range(1, number_of_scenarios + 1)]
    mean_pv = sum(pv_values) / len(pv_values)
    dashboard_results = {
        "Mean_PV_Cash_Flow": mean_pv,
        "Median_PV_Cash_Flow": sorted(pv_values)[len(pv_values) // 2],
        "Std_Dev_PV_Cash_Flow": (sum((x - mean_pv) ** 2 for x in pv_values) / len(pv_values)) ** 0.5,
        "Min_PV_Cash_Flow": min(pv_values),
        "Max_PV_Cash_Flow": max(pv_values),
    }
    total_runtime_seconds = int(runtime_seconds)
    runtime_formatted = (
        f"{total_runtime_seconds // 3600:02d}:"
        f"{(total_runtime_seconds % 3600) // 60:02d}:"
        f"{total_runtime_seconds % 60:02d}"
    )

    worksheet = workbook.create_sheet("Dashboard Results")
    worksheet.append(["Number of Scenarios", None, number_of_scenarios])
    worksheet.append(["Number of Policies", None, number_of_policies])
    worksheet.append(["Discount rate", None, discount_rate])
    worksheet.append(["PV Cash Flow Statistics"])
    worksheet.append(["Mean", "Median", "Std Dev", "Min", "Max", "Runtime"])
    worksheet.append(
        [
            dashboard_results["Mean_PV_Cash_Flow"],
            dashboard_results["Median_PV_Cash_Flow"],
            dashboard_results["Std_Dev_PV_Cash_Flow"],
            dashboard_results["Min_PV_Cash_Flow"],
            dashboard_results["Max_PV_Cash_Flow"],
            runtime_formatted,
        ]
    )

    if str(report_value("Dashboard Results", "Create graph?", "No")).strip().lower() == "yes":
        figure = make_dashboard_results_figure(number_of_scenarios, total_results_by_scenario, pv_cash_flows_by_scenario)
        if figure is not None:
            embed_chart(worksheet, figure, "A" + str(worksheet.max_row + 2))


def make_scenario_results_figure(
    scenario_number: int,
    number_of_policies: int,
    scenario_results: list[dict[str, int]],
):
    years = [row["Year"] for row in scenario_results]
    total_cash_flows = [row["Total"] for row in scenario_results]
    figure, chart = plt.subplots(figsize=(10, 5))

    if number_of_policies <= 10:
        bottom = [0] * len(years)
        for policy in Inforce:
            policy_cash_flows = [row[f"Policy_{policy.Policy}"] for row in scenario_results]
            chart.bar(years, policy_cash_flows, bottom=bottom, label=f"Policy_{policy.Policy}")
            bottom = [bottom[i] + policy_cash_flows[i] for i in range(len(years))]

    chart.plot(years, total_cash_flows, color="black", marker="o", linestyle="--", label="Total")
    chart.set_xlabel("Year")
    chart.set_ylabel("Cash Flow")
    chart.set_title(f"Scenario {scenario_number} Cash Flows by Policy")
    chart.legend()
    figure.tight_layout()
    return figure


def make_dashboard_results_figure(
    number_of_scenarios: int,
    total_results_by_scenario: list[dict[str, int]],
    pv_cash_flows_by_scenario: dict[int, float] | None = None,
):
    years = [row["Year"] for row in total_results_by_scenario]
    figure, chart = plt.subplots(figsize=(10, 5))

    if number_of_scenarios > 25 and pv_cash_flows_by_scenario is not None:
        pv_values = np.array([pv_cash_flows_by_scenario[i + 1] for i in range(number_of_scenarios)])
        sorted_indices = np.argsort(pv_values)
        positions = np.round(np.linspace(0, number_of_scenarios - 1, 11)).astype(int)
        for pct_idx, scen_idx in enumerate(sorted_indices[positions]):
            scen_num = int(scen_idx) + 1
            cash_flows = [row[f"Scenario_{scen_num}"] for row in total_results_by_scenario]
            chart.plot(years, cash_flows, label=f"Percentile_{pct_idx * 10}_Scenario_{scen_num}")
    else:
        for scenario_number in range(1, number_of_scenarios + 1):
            cash_flows = [row[f"Scenario_{scenario_number}"] for row in total_results_by_scenario]
            chart.plot(years, cash_flows, label=f"Scenario_{scenario_number}")

    chart.set_xlabel("Year")
    chart.set_ylabel("Cash Flow")
    chart.set_title("Cash Flow Projection by Scenario")
    chart.legend()
    figure.tight_layout()
    return figure


def embed_chart(worksheet, figure, anchor: str) -> None:
    image_buffer = io.BytesIO()
    figure.savefig(image_buffer, format="png", dpi=100)
    plt.close(figure)
    image_buffer.seek(0)
    worksheet.add_image(XlImage(image_buffer), anchor)

# ==========================================================
# Main execution function to run the model and print results
# ==========================================================

if __name__ == "__main__":

    start_time = time.time()

    if len(sys.argv) != 2:
        print('Usage: python RnD_model.py "inputs/Input 10 pol 25 scen table.xlsx"')
        sys.exit(1)

    input_file_path = Path(sys.argv[1])
    if not input_file_path.exists():
        print(f"Error: input file not found: {input_file_path}")
        sys.exit(1)

    print("R&D model started.")
    print(f"Input file: {input_file_path}")

    try:
        load_input_workbook(input_file_path)
        output_file_path = generate_output_xlsx(start_time)
    except Exception as error:
        print(f"Error: {error}")
        sys.exit(1)
    
    end_time = time.time()
    print(f"Output file: {output_file_path}")
    print(f"Runtime: {end_time - start_time:.2f} seconds")
