"""Unit tests for ReportWriter."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import numpy as np
import openpyxl
import pytest

from mbc_model.data.models import (
    ModelResults,
    PolicyDetail,
    PolicyRecord,
    ReportingConfig,
)
from mbc_model.reporting.writer import ReportWriter


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _policy(pid: int, yob: int, gender: str) -> PolicyRecord:
    return PolicyRecord(policy_id=pid, yob=yob, gender=gender, annual_benefit=1000.0)  # type: ignore[arg-type]


def _make_results(
    create_policy: bool = False,
    create_scenario: bool = False,
    create_total: bool = False,
    create_dashboard: bool = False,
    create_scenario_graph: bool = False,
    create_dashboard_graph: bool = False,
) -> ModelResults:
    n_policies = 2
    n_years = 5
    n_scenarios = 3
    years = np.arange(2025, 2030, dtype=np.int64)
    cum_survival = np.full((n_policies, n_years), 0.9, dtype=np.float64)
    scenario_cfs = np.full((n_scenarios, n_years), 500.0, dtype=np.float64)
    pv = np.array([1000.0, 900.0, 800.0], dtype=np.float64)

    detail: PolicyDetail | None = None
    if create_policy:
        detail = PolicyDetail(
            policy_id=1,
            scenario_id=1,
            projection_years=years.copy(),
            ages=np.array([74, 75, 76, 77, 78], dtype=np.int64),
            base_qx=np.full(5, 0.05),
            improvement=np.full(5, 0.99),
            improved_qx=np.full(5, 0.0495),
            px=np.full(5, 0.9505),
            cum_survival=np.cumprod(np.full(5, 0.9505)),
            cum_death=1.0 - np.cumprod(np.full(5, 0.9505)),
            survive_flag=np.ones(5),
            annual_cf=np.full(5, 1000.0),
        )

    cfg = ReportingConfig(
        create_policy_results=create_policy,
        policy_id=1,
        policy_scenario_id=1,
        create_scenario_results=create_scenario,
        scenario_id=1,
        create_scenario_graph=create_scenario_graph,
        create_total_results=create_total,
        create_dashboard_results=create_dashboard,
        dashboard_scenarios=n_scenarios,
        dashboard_contracts=n_policies,
        discount_rate=0.04,
        create_dashboard_graph=create_dashboard_graph,
    )
    return ModelResults(
        policies=[_policy(1, 1951, "M"), _policy(2, 1950, "F")],
        projection_years=years,
        cum_survival=cum_survival,
        scenario_cash_flows=scenario_cfs,
        pv_by_scenario=pv if create_dashboard else None,
        policy_detail=detail,
        runtime_seconds=0.5,
        config=cfg,
    )


# ---------------------------------------------------------------------------
# File creation
# ---------------------------------------------------------------------------


def test_write_creates_file(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(_make_results())
    assert out.exists()
    assert out.suffix == ".xlsx"


def test_write_creates_output_dir_if_missing(tmp_path: Path) -> None:
    out_dir = tmp_path / "subdir" / "nested"
    out = ReportWriter(out_dir).write(_make_results())
    assert out.exists()


def test_run_summary_always_written(tmp_path: Path) -> None:
    """Run Summary sheet is always present so the workbook is never empty."""
    out = ReportWriter(tmp_path).write(_make_results())
    wb = openpyxl.load_workbook(out)
    assert "Run Summary" in wb.sheetnames


# ---------------------------------------------------------------------------
# Policy Results sheet
# ---------------------------------------------------------------------------


def test_policy_results_sheet_created(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(_make_results(create_policy=True))
    wb = openpyxl.load_workbook(out)
    assert "Policy Results" in wb.sheetnames


def test_policy_results_row_count(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(_make_results(create_policy=True))
    ws = openpyxl.load_workbook(out)["Policy Results"]
    assert ws.max_row == 6  # 1 header + 5 data rows


def test_policy_results_header(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(_make_results(create_policy=True))
    ws = openpyxl.load_workbook(out)["Policy Results"]
    header = [ws.cell(1, c).value for c in range(1, 11)]
    assert header[0] == "Year"
    assert header[-1] == "Cash_Flow"


# ---------------------------------------------------------------------------
# Scenario Results sheet
# ---------------------------------------------------------------------------


def test_scenario_results_sheet_created(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(_make_results(create_scenario=True))
    wb = openpyxl.load_workbook(out)
    assert "Scenario Results" in wb.sheetnames


def test_scenario_results_row_count(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(_make_results(create_scenario=True))
    ws = openpyxl.load_workbook(out)["Scenario Results"]
    assert ws.max_row == 6  # 1 header + 5 years


# ---------------------------------------------------------------------------
# Total Results sheet
# ---------------------------------------------------------------------------


def test_total_results_sheet_created(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(_make_results(create_total=True))
    wb = openpyxl.load_workbook(out)
    assert "Total Results" in wb.sheetnames


def test_total_results_row_count(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(_make_results(create_total=True))
    ws = openpyxl.load_workbook(out)["Total Results"]
    assert ws.max_row == 6  # 1 header + 5 years


def test_total_results_values(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(_make_results(create_total=True))
    ws = openpyxl.load_workbook(out)["Total Results"]
    assert ws.cell(2, 1).value == 2025  # first year
    assert ws.cell(2, 2).value == pytest.approx(500.0)  # scenario 1, year 1


# ---------------------------------------------------------------------------
# Dashboard Results sheet
# ---------------------------------------------------------------------------


def test_dashboard_results_sheet_created(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(_make_results(create_dashboard=True))
    wb = openpyxl.load_workbook(out)
    assert "Dashboard Results" in wb.sheetnames


def test_dashboard_results_mean_pv(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(_make_results(create_dashboard=True))
    ws = openpyxl.load_workbook(out)["Dashboard Results"]
    mean_row = None
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Mean PV Cash Flow":
            mean_row = row
            break
    assert mean_row is not None
    assert mean_row[1] == pytest.approx(900.0)  # mean of [1000, 900, 800]


# ---------------------------------------------------------------------------
# Chart embedding (smoke tests — just check no exception raised)
# ---------------------------------------------------------------------------


def test_scenario_graph_no_error(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(
        _make_results(create_scenario=True, create_scenario_graph=True)
    )
    assert out.exists()


def test_dashboard_graph_no_error(tmp_path: Path) -> None:
    out = ReportWriter(tmp_path).write(
        _make_results(create_dashboard=True, create_dashboard_graph=True)
    )
    assert out.exists()


def test_dashboard_skipped_when_pv_none(tmp_path: Path) -> None:
    """If create_dashboard_results=True but pv_by_scenario is None, sheet is skipped."""
    results = _make_results(create_dashboard=True)
    results = ModelResults(
        policies=results.policies,
        projection_years=results.projection_years,
        cum_survival=results.cum_survival,
        scenario_cash_flows=results.scenario_cash_flows,
        pv_by_scenario=None,  # override to None
        policy_detail=results.policy_detail,
        runtime_seconds=results.runtime_seconds,
        config=results.config,
    )
    out = ReportWriter(tmp_path).write(results)
    wb = openpyxl.load_workbook(out)
    assert "Dashboard Results" not in wb.sheetnames
