"""Tests for validated desktop input workbook editing."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from mbc_model.data.loader import ExcelLoader
from mbc_model.ui.input_editor import InputEditError, save_input_copy, validate_input_changes

_ROOT = Path(__file__).parent.parent.parent
_SMALL_TABLE = _ROOT / "inputs" / "Input 10 pol 25 scen table.xlsx"


@pytest.mark.skipif(not _SMALL_TABLE.exists(), reason=f"Fixture workbook not found: {_SMALL_TABLE}")
def test_save_input_copy_updates_allowed_edits(tmp_path: Path) -> None:
    source_path = _make_xlsx_fixture(tmp_path)
    output_path = tmp_path / "edited.xlsx"
    changes = {
        "inforce": [{"policy_id": 1, "field": "Annual Benefit", "value": "21000"}],
        "parameters": {
            "projection_settings": [
                {"parameter": "Random Numbers Seed", "value": "42"},
            ],
            "mortality_rates": [{"age": 0, "male": "0.001"}],
            "projection_scale": [{"age": 0, "female": "0.02"}],
            "random_numbers": [{"scenario": 1, "policy": 1, "value": "0.5"}],
        },
        "reporting": [{"section": "Dashboard Results", "field": "Discount Rate", "value": "0.05"}],
    }

    saved_path = save_input_copy(source_path, output_path, changes)

    assert saved_path == output_path
    policies = ExcelLoader(output_path).load_inforce()
    parameters, mortality, random_numbers = ExcelLoader(output_path).load_parameters()
    reporting = ExcelLoader(output_path).load_reporting()
    assert policies[0].annual_benefit == 21000
    assert parameters.random_seed == 42
    assert mortality.male_mortality_rates[0] == pytest.approx(0.001)
    assert mortality.female_projection_scale[0] == pytest.approx(0.02)
    assert random_numbers is not None
    assert random_numbers[0, 0] == pytest.approx(0.5)
    assert reporting.discount_rate == pytest.approx(0.05)


@pytest.mark.skipif(not _SMALL_TABLE.exists(), reason=f"Fixture workbook not found: {_SMALL_TABLE}")
def test_save_input_copy_rejects_unsafe_save_targets_and_invalid_edits(tmp_path: Path) -> None:
    source_path = _make_xlsx_fixture(tmp_path)

    with pytest.raises(InputEditError, match="different file name"):
        save_input_copy(source_path, source_path, {})

    with pytest.raises(InputEditError, match=".xlsx"):
        save_input_copy(source_path, tmp_path / "edited.xlsm", {})

    with pytest.raises(InputEditError, match="YOB"):
        save_input_copy(
            source_path,
            tmp_path / "edited.xlsx",
            {"inforce": [{"policy_id": 1, "field": "YOB", "value": "1700"}]},
        )


@pytest.mark.skipif(not _SMALL_TABLE.exists(), reason=f"Fixture workbook not found: {_SMALL_TABLE}")
def test_validate_input_changes_rejects_unknown_field(tmp_path: Path) -> None:
    source_path = _make_xlsx_fixture(tmp_path)
    result = validate_input_changes(
        source_path,
        {"inforce": [{"policy_id": 1, "field": "Policy #", "value": "2"}]},
    )

    assert not result.ok
    assert "not editable" in result.errors[0]


@pytest.mark.skipif(not _SMALL_TABLE.exists(), reason=f"Fixture workbook not found: {_SMALL_TABLE}")
def test_validate_input_changes_reports_inforce_value_errors(tmp_path: Path) -> None:
    source_path = _make_xlsx_fixture(tmp_path)

    result = validate_input_changes(
        source_path,
        {
            "inforce": [
                {"policy_id": "bad", "field": "YOB", "value": "1950"},
                {"policy_id": 999, "field": "YOB", "value": "1950"},
                {"policy_id": 1, "field": "Gender", "value": "X"},
                {"policy_id": 1, "field": "Annual Benefit", "value": "-1"},
            ]
        },
    )

    assert not result.ok
    assert any("must be an integer" in error for error in result.errors)
    assert any("not in the workbook" in error for error in result.errors)
    assert any("Gender must be M or F" in error for error in result.errors)
    assert any("Annual Benefit" in error for error in result.errors)


@pytest.mark.skipif(not _SMALL_TABLE.exists(), reason=f"Fixture workbook not found: {_SMALL_TABLE}")
def test_validate_input_changes_reports_missing_inforce_column(tmp_path: Path) -> None:
    source_path = tmp_path / "missing-column.xlsx"
    workbook = openpyxl.load_workbook(_SMALL_TABLE)
    workbook["Inforce"].cell(1, 2).value = "Missing YOB"
    workbook.save(source_path)

    result = validate_input_changes(
        source_path,
        {"inforce": [{"policy_id": 1, "field": "YOB", "value": "1955"}]},
    )

    assert not result.ok
    assert "was not found" in result.errors[0]


@pytest.mark.skipif(not _SMALL_TABLE.exists(), reason=f"Fixture workbook not found: {_SMALL_TABLE}")
def test_validate_input_changes_rejects_formula_cell(tmp_path: Path) -> None:
    formula_path = tmp_path / "formula.xlsx"
    workbook = openpyxl.load_workbook(_SMALL_TABLE)
    workbook["Inforce"].cell(2, 4).value = "=1+1"
    workbook.save(formula_path)

    result = validate_input_changes(
        formula_path,
        {"inforce": [{"policy_id": 1, "field": "Annual Benefit", "value": "21000"}]},
    )

    assert not result.ok
    assert "formula cell" in result.errors[0]


@pytest.mark.skipif(not _SMALL_TABLE.exists(), reason=f"Fixture workbook not found: {_SMALL_TABLE}")
def test_validate_input_changes_rejects_large_inforce_edits(tmp_path: Path) -> None:
    large_path = tmp_path / "large.xlsx"
    workbook = openpyxl.load_workbook(_SMALL_TABLE)
    inforce = workbook["Inforce"]
    for policy_id in range(11, 102):
        row = policy_id + 1
        inforce.cell(row, 1).value = policy_id
        inforce.cell(row, 2).value = 1950
        inforce.cell(row, 3).value = "M"
        inforce.cell(row, 4).value = 10000
    workbook.save(large_path)

    result = validate_input_changes(
        large_path,
        {"inforce": [{"policy_id": 1, "field": "YOB", "value": "1955"}]},
    )

    assert not result.ok
    assert "100 policies or fewer" in result.errors[0]


@pytest.mark.skipif(not _SMALL_TABLE.exists(), reason=f"Fixture workbook not found: {_SMALL_TABLE}")
def test_validate_input_changes_reports_parameter_errors(tmp_path: Path) -> None:
    source_path = _make_xlsx_fixture(tmp_path)

    result = validate_input_changes(
        source_path,
        {
            "parameters": {
                "projection_settings": [
                    {"parameter": "Unknown Parameter", "value": "1"},
                    {"parameter": "Valuation Date", "value": "not-a-date"},
                    {"parameter": "Which Random Numbers?", "value": "Maybe"},
                ],
                "mortality_rates": [
                    {"age": 999, "male": "0.01"},
                    {"age": 0, "female": "2.0"},
                ],
                "projection_scale": [{"age": "bad", "male": "0.01"}],
            }
        },
    )

    assert not result.ok
    assert any("not editable" in error for error in result.errors)
    assert any("valid date" in error for error in result.errors)
    assert any("Seed or Table" in error for error in result.errors)
    assert any("outside the workbook table" in error for error in result.errors)
    assert any("between 0 and 1" in error for error in result.errors)
    assert any("must be an integer" in error for error in result.errors)


@pytest.mark.skipif(not _SMALL_TABLE.exists(), reason=f"Fixture workbook not found: {_SMALL_TABLE}")
def test_validate_input_changes_reports_missing_parameter_table(tmp_path: Path) -> None:
    source_path = tmp_path / "missing-marker.xlsx"
    workbook = openpyxl.load_workbook(_SMALL_TABLE)
    parameters = workbook["Parameters"]
    for row in range(1, parameters.max_row + 1):
        if parameters.cell(row, 1).value == "Mortality Rates":
            parameters.cell(row, 1).value = "Missing Mortality Rates"
            break
    workbook.save(source_path)

    result = validate_input_changes(
        source_path,
        {"parameters": {"mortality_rates": [{"age": 0, "male": "0.01"}]}},
    )

    assert not result.ok
    assert "table was not found" in result.errors[0]


@pytest.mark.skipif(not _SMALL_TABLE.exists(), reason=f"Fixture workbook not found: {_SMALL_TABLE}")
def test_validate_input_changes_reports_random_number_and_reporting_errors(
    tmp_path: Path,
) -> None:
    seed_path = _ROOT / "inputs" / "Input 10 pol 25 scen seed.xlsx"
    source_path = _make_xlsx_fixture(tmp_path)

    seed_result = validate_input_changes(
        seed_path,
        {"parameters": {"random_numbers": [{"scenario": 1, "policy": 1, "value": "0.5"}]}},
    )
    assert not seed_result.ok
    assert "random number table" in seed_result.errors[0]

    result = validate_input_changes(
        source_path,
        {
            "parameters": {
                "random_numbers": [
                    {"scenario": 999, "policy": 1, "value": "0.5"},
                    {"scenario": 1, "policy": 1, "value": "bad"},
                ]
            },
            "reporting": [
                {"section": "Unknown", "field": "Create?", "value": "Yes"},
                {"section": "Dashboard Results", "field": "Create?", "value": "Maybe"},
            ],
        },
    )

    assert not result.ok
    assert any("outside the table" in error for error in result.errors)
    assert any("must be numeric" in error for error in result.errors)
    assert any("not editable" in error for error in result.errors)
    assert any("must be Yes or No" in error for error in result.errors)


def _make_xlsx_fixture(tmp_path: Path) -> Path:
    output_path = tmp_path / "input.xlsx"
    workbook = openpyxl.load_workbook(_SMALL_TABLE)
    workbook.save(output_path)
    return output_path
