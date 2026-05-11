"""Tests for UI workbook preview payloads."""

from __future__ import annotations

import base64
import os
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook

from mbc_model.data.models import ModelResults
from mbc_model.reporting.charts import GRAPH_NOT_REQUESTED_MESSAGE
from mbc_model.runner import run, run_with_results
from mbc_model.ui.preview import PREVIEW_LIMIT, build_input_preview, build_output_preview

_ROOT = Path(__file__).parent.parent.parent
_SMALL_SEED = _ROOT / "inputs" / "Input 10 pol 25 scen seed.xlsm"
_LARGE_SEED = _ROOT / "inputs" / "Input 50k pol 10k scen seed.xlsm"


@pytest.mark.skipif(not _SMALL_SEED.exists(), reason=f"Fixture workbook not found: {_SMALL_SEED}")
def test_input_preview_small_seed_workbook_counts_actual_records() -> None:
    preview = build_input_preview(_SMALL_SEED)

    assert preview["policy_count"] == 10
    assert preview["scenario_count"] == 25
    assert preview["random_numbers"] == "Seed"
    assert preview["inforce"]["shown_rows"] == 10
    assert not preview["inforce"]["has_more_rows"]


@pytest.mark.skipif(not _LARGE_SEED.exists(), reason=f"Fixture workbook not found: {_LARGE_SEED}")
def test_input_preview_large_seed_workbook_is_capped() -> None:
    preview = build_input_preview(_LARGE_SEED)

    assert preview["policy_count"] == 50000
    assert preview["scenario_count"] == 10000
    assert preview["inforce"]["shown_rows"] == PREVIEW_LIMIT
    assert preview["inforce"]["has_more_rows"]
    assert preview["parameters"]["random_numbers"]["rows"][0] == [
        "Which Random Numbers?",
        "Seed",
    ]


@pytest.mark.skipif(not _SMALL_SEED.exists(), reason=f"Fixture workbook not found: {_SMALL_SEED}")
def test_output_preview_reads_small_run_results(tmp_path: Path) -> None:
    output_path = run(_SMALL_SEED, output_dir=tmp_path, verbose=False)
    preview = build_output_preview(output_path)

    assert preview["file_name"].startswith("results_")
    assert preview["dashboard"]["available"]
    assert preview["total"]["available"]
    assert preview["scenario"]["available"]
    assert preview["policy"]["available"]
    assert preview["dashboard"]["rows"][0] == ["Discount rate", "4.00%"]
    assert preview["total"]["cash_flow"]["shown_rows"] <= PREVIEW_LIMIT


@pytest.mark.skipif(not _SMALL_SEED.exists(), reason=f"Fixture workbook not found: {_SMALL_SEED}")
def test_output_preview_includes_real_chart_data_urls(tmp_path: Path) -> None:
    output_path, results = run_with_results(_SMALL_SEED, output_dir=tmp_path, verbose=False)

    preview = build_output_preview(output_path, results)

    _assert_png_data_url(preview["scenario"]["chart"]["data_url"])
    _assert_png_data_url(preview["dashboard"]["chart"]["data_url"])
    assert preview["scenario"]["chart"]["title"] == "Scenario 25 Cash Flows by Policy"
    assert preview["dashboard"]["chart"]["title"] == "Cash Flow Projection by Scenario"


@pytest.mark.skipif(not _SMALL_SEED.exists(), reason=f"Fixture workbook not found: {_SMALL_SEED}")
def test_output_preview_says_when_graphs_were_not_requested(tmp_path: Path) -> None:
    output_path, results = run_with_results(_SMALL_SEED, output_dir=tmp_path, verbose=False)
    disabled_results = _with_graph_flags(results, scenario=False, dashboard=False)

    preview = build_output_preview(output_path, disabled_results)

    assert not preview["scenario"]["chart"]["available"]
    assert not preview["dashboard"]["chart"]["available"]
    assert preview["scenario"]["chart"]["message"] == GRAPH_NOT_REQUESTED_MESSAGE
    assert preview["dashboard"]["chart"]["message"] == GRAPH_NOT_REQUESTED_MESSAGE


def test_output_preview_flags_more_pv_cash_flow_scenarios(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Total Results"
    sheet.cell(1, 1, "Discount Rate")
    sheet.cell(1, 2, 0.04)
    sheet.cell(2, 1, "PV Cash Flow")
    for column in range(2, 103):
        sheet.cell(2, column, column * 100)
    path = tmp_path / "results.xlsx"
    workbook.save(path)

    preview = build_output_preview(path)

    assert preview["total"]["pv_cash_flow"]["shown_columns"] == PREVIEW_LIMIT
    assert preview["total"]["pv_cash_flow"]["has_more_columns"]


@pytest.mark.slow
@pytest.mark.skipif(
    os.getenv("RUN_SLOW_UI_TESTS") != "1",
    reason="Set RUN_SLOW_UI_TESTS=1 to run the 50k policy acceptance test.",
)
@pytest.mark.skipif(not _LARGE_SEED.exists(), reason=f"Fixture workbook not found: {_LARGE_SEED}")
def test_large_seed_workbook_graph_payload_acceptance(tmp_path: Path) -> None:
    output_path, results = run_with_results(_LARGE_SEED, output_dir=tmp_path, verbose=False)

    preview = build_output_preview(output_path, results)

    _assert_png_data_url(preview["scenario"]["chart"]["data_url"])
    _assert_png_data_url(preview["dashboard"]["chart"]["data_url"])
    assert preview["total"]["cash_flow"]["shown_rows"] <= PREVIEW_LIMIT


def _with_graph_flags(
    results: ModelResults,
    scenario: bool,
    dashboard: bool,
) -> ModelResults:
    return replace(
        results,
        config=replace(
            results.config,
            create_scenario_graph=scenario,
            create_dashboard_graph=dashboard,
        ),
    )


def _assert_png_data_url(data_url: str) -> None:
    prefix = "data:image/png;base64,"
    assert data_url.startswith(prefix)
    image_bytes = base64.b64decode(data_url.removeprefix(prefix))
    assert image_bytes.startswith(b"\x89PNG\r\n\x1a\n")
    assert len(image_bytes) > 1000
